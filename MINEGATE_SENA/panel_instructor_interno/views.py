from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.db.models import Q
from django.db import IntegrityError
from django.db import transaction
import os
import csv
import io
from visitaInterna.models import VisitaInterna
from .models import Ficha, Programa
from .forms import VisitaInternaInstructorForm, ProgramaForm, FichaForm
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from calendario.models import ReservaHorario

from django.http import JsonResponse
from visitaInterna.models import (
    HistorialAccionVisitaInterna,
    HistorialReprogramacion,
    VisitaInterna,
    AsistenteVisitaInterna,
)
from .models import Ficha, Programa, Aprendiz
from .forms import VisitaInternaInstructorForm, ProgramaForm, FichaForm, AprendizForm
from documentos.models import Documento, DocumentoSubidoAsistente
from documentos.models import (
    Documento,
    DocumentoSubidoAprendiz,
    DocumentoSubidoAsistente,
)

CATEGORIA_DOC_SALUD = "Formato Auto Reporte Condiciones de Salud"
CATEGORIA_AUTORIZACION_PADRES = "Formato Autorización Padres de Familia"
CATEGORIAS_ARCHIVOS_FINALES = [
    "ATS",
    "Formato Inducción y Reinducción",
    "Charla de Seguridad y Calestenia",
]
MARCADOR_SOLICITUD_FINAL_ENVIADA = "[SOLICITUD_FINAL_ENVIADA]"
EXTENSIONES_DOCUMENTOS_APRENDIZ = {".pdf", ".doc", ".docx"}
MAX_TAMANO_DOCUMENTO_APRENDIZ = 10 * 1024 * 1024
FORMATOS_CARGA_MASIVA_APRENDICES = {".csv", ".xlsx", ".xls", ".pdf"}
COLUMNAS_CARGA_MASIVA_APRENDICES = [
    "nombre",
    "apellido",
    "tipo_documento",
    "numero_documento",
    "correo",
    "telefono",
    "estado",
]
EJEMPLOS_CARGA_MASIVA_APRENDICES = [
    {
        "nombre": "Laura",
        "apellido": "Gomez",
        "tipo_documento": "CC",
        "numero_documento": "1022334455",
        "correo": "laura.gomez@example.com",
        "telefono": "3001234567",
        "estado": "activo",
    },
    {
        "nombre": "Mateo",
        "apellido": "Rodriguez",
        "tipo_documento": "TI",
        "numero_documento": "1099988877",
        "correo": "mateo.rodriguez@example.com",
        "telefono": "3017654321",
        "estado": "activo",
    },
]


def validar_archivos_documentos_aprendiz(files):
    """Valida que los documentos dinámicos del aprendiz solo sean PDF o Word."""
    errores = []

    for campo, archivo in files.items():
        if not campo.startswith("documento_") or not archivo:
            continue

        extension = os.path.splitext(archivo.name)[1].lower()
        if extension not in EXTENSIONES_DOCUMENTOS_APRENDIZ:
            errores.append(
                f'❌ El archivo "{archivo.name}" no es válido. Solo se permiten PDF o Word (.doc, .docx).'
            )

    return errores


def _normalizar_encabezado_importacion(valor):
    return (
        str(valor or "")
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )


def _mapear_tipo_documento_importacion(valor):
    tipo = str(valor or "").strip().upper()
    equivalencias = {
        "CEDULA": "CC",
        "CEDULA_DE_CIUDADANIA": "CC",
        "TARJETA_DE_IDENTIDAD": "TI",
        "PASAPORTE": "PP",
    }
    return equivalencias.get(tipo, tipo)


def _mapear_estado_aprendiz_importacion(valor):
    estado = _normalizar_encabezado_importacion(valor).replace("_", "")
    equivalencias = {
        "": "activo",
        "activo": "activo",
        "inactivo": "inactivo",
        "retirado": "retirado",
    }
    return equivalencias.get(estado, "activo")


def _extraer_filas_csv_aprendices(archivo):
    contenido = archivo.read()
    if isinstance(contenido, bytes):
        try:
            texto = contenido.decode("utf-8-sig")
        except UnicodeDecodeError:
            texto = contenido.decode("latin-1")
    else:
        texto = str(contenido)

    lector = csv.DictReader(io.StringIO(texto))
    if not lector.fieldnames:
        return [], ["El CSV no contiene encabezados."]

    headers = [_normalizar_encabezado_importacion(h) for h in lector.fieldnames]
    faltantes = [c for c in COLUMNAS_CARGA_MASIVA_APRENDICES if c not in headers]
    if faltantes:
        return [], [
            "La plantilla no coincide. Faltan columnas: " + ", ".join(faltantes)
        ]

    filas = []
    errores = []
    for idx, row in enumerate(lector, start=2):
        fila = {}
        for key, value in row.items():
            fila[_normalizar_encabezado_importacion(key)] = (value or "").strip()
        if not any(fila.values()):
            continue
        filas.append((idx, fila))

    if not filas and not errores:
        errores.append("No se encontraron filas de aprendices para importar.")

    return filas, errores


def _extraer_filas_excel_aprendices(archivo):
    try:
        from openpyxl import load_workbook
    except Exception:
        return [], [
            "No se pudo procesar Excel. Instale openpyxl para habilitar este formato."
        ]

    try:
        wb = load_workbook(filename=archivo, data_only=True)
        ws = wb.active
    except Exception:
        return [], ["No fue posible leer el archivo Excel. Verifique la plantilla."]

    encabezados = []
    for cell in ws[1]:
        encabezados.append(_normalizar_encabezado_importacion(cell.value))

    faltantes = [c for c in COLUMNAS_CARGA_MASIVA_APRENDICES if c not in encabezados]
    if faltantes:
        return [], [
            "La plantilla no coincide. Faltan columnas: " + ", ".join(faltantes)
        ]

    filas = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        fila = {}
        for col_idx, valor in enumerate(row):
            if col_idx >= len(encabezados):
                continue
            clave = encabezados[col_idx]
            fila[clave] = "" if valor is None else str(valor).strip()

        if not any(fila.values()):
            continue
        filas.append((idx, fila))

    if not filas:
        return [], ["No se encontraron filas de aprendices para importar."]

    return filas, []


def _extraer_filas_pdf_aprendices(archivo):
    try:
        from pypdf import PdfReader
    except Exception:
        return [], ["No se pudo procesar PDF. Instale pypdf para habilitar este formato."]

    try:
        reader = PdfReader(archivo)
        texto = "\n".join((p.extract_text() or "") for p in reader.pages)
    except Exception:
        return [], ["No fue posible leer el archivo PDF. Use la plantilla descargable."]

    filas = []
    for idx, linea in enumerate(texto.splitlines(), start=1):
        normalizada = linea.strip()
        if not normalizada or ";" not in normalizada:
            continue
        partes = [p.strip() for p in normalizada.split(";")]
        if len(partes) < 7:
            continue
        if _normalizar_encabezado_importacion(partes[0]) == "nombre":
            continue

        fila = {
            "nombre": partes[0],
            "apellido": partes[1],
            "tipo_documento": partes[2],
            "numero_documento": partes[3],
            "correo": partes[4],
            "telefono": partes[5],
            "estado": partes[6],
        }
        filas.append((idx, fila))

    if not filas:
        return [], [
            "No se detectaron filas válidas en el PDF. Use la plantilla y mantenga el formato con ';'."
        ]

    return filas, []


def _extraer_filas_importacion_aprendices(archivo):
    extension = os.path.splitext(getattr(archivo, "name", ""))[1].lower()

    if extension not in FORMATOS_CARGA_MASIVA_APRENDICES:
        return [], ["Formato no permitido. Use CSV, Excel (.xlsx/.xls) o PDF."]

    if extension == ".csv":
        return _extraer_filas_csv_aprendices(archivo)

    if extension in {".xlsx", ".xls"}:
        return _extraer_filas_excel_aprendices(archivo)

    return _extraer_filas_pdf_aprendices(archivo)


def _procesar_carga_masiva_aprendices(ficha, filas):
    creados = 0
    duplicados = 0
    omitidos_cupo = 0
    errores = []
    cupos_disponibles = max(0, ficha.cantidad_aprendices - ficha.aprendices.count())

    for fila_numero, fila in filas:
        if creados >= cupos_disponibles:
            omitidos_cupo += 1
            continue

        datos_form = {
            "nombre": fila.get("nombre", ""),
            "apellido": fila.get("apellido", ""),
            "tipo_documento": _mapear_tipo_documento_importacion(
                fila.get("tipo_documento")
            ),
            "numero_documento": str(fila.get("numero_documento", "")).strip(),
            "correo": str(fila.get("correo", "")).strip().lower(),
            "telefono": str(fila.get("telefono", "")).strip(),
            "estado": _mapear_estado_aprendiz_importacion(fila.get("estado")),
        }

        if not datos_form["numero_documento"]:
            errores.append(f"Fila {fila_numero}: el numero de documento es obligatorio.")
            continue

        existe = Aprendiz.objects.filter(
            ficha=ficha,
            numero_documento=datos_form["numero_documento"],
        ).exists()
        if existe:
            duplicados += 1
            continue

        form = AprendizForm(data=datos_form, ficha=ficha)
        if not form.is_valid():
            errores_fila = []
            for campo, errs in form.errors.items():
                nombre_campo = "general" if campo == "__all__" else campo
                errores_fila.append(f"{nombre_campo}: {', '.join(errs)}")
            errores.append(f"Fila {fila_numero}: {' | '.join(errores_fila)}")
            continue

        aprendiz = form.save(commit=False)
        aprendiz.ficha = ficha
        aprendiz.save()
        creados += 1

    return {
        "creados": creados,
        "duplicados": duplicados,
        "omitidos_cupo": omitidos_cupo,
        "errores": errores,
    }


def _normalizar_categoria_texto(value):
    return (
        str(value or "")
        .lower()
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )


def _es_categoria_archivo_final(categoria):
    cat = _normalizar_categoria_texto(categoria)
    if "ats" in cat:
        return True
    if "induccion y reinduccion" in cat:
        return True
    return "charla de seguridad" in cat and ("calestenia" in cat or "calistenia" in cat)


def construir_reporte_documental_visita(visita, tipo_visita):
    """Genera un resumen de faltantes y rechazos por asistente y archivos finales."""
    doc_salud_ids = set(
        Documento.objects.filter(categoria=CATEGORIA_DOC_SALUD).values_list(
            "id", flat=True
        )
    )
    docs_finales_requeridos = list(
        Documento.objects.filter(categoria__in=CATEGORIAS_ARCHIVOS_FINALES).order_by(
            "categoria", "titulo"
        )
    )

    filtros_finales_visita = {
        "documento_requerido__categoria__in": CATEGORIAS_ARCHIVOS_FINALES
    }
    if tipo_visita == "interna":
        filtros_finales_visita["asistente_interna__visita"] = visita
    else:
        filtros_finales_visita["asistente_externa__visita"] = visita

    existe_archivo_final_subido = DocumentoSubidoAsistente.objects.filter(
        **filtros_finales_visita
    ).exists()
    mostrar_faltantes_finales_como_alerta = (
        visita.estado in ["documentos_enviados", "en_revision_documentos", "confirmada"]
        or existe_archivo_final_subido
    )

    asistentes_con_alertas = []
    asistentes = visita.asistentes.prefetch_related(
        "documentos_subidos__documento_requerido"
    )
    for asistente in asistentes:
        incidencias = []
        latest_por_documento = {}
        for ds in asistente.documentos_subidos.all():
            doc_id = ds.documento_requerido_id
            actual = latest_por_documento.get(doc_id)
            if not actual or (ds.fecha_subida, ds.id) > (
                actual.fecha_subida,
                actual.id,
            ):
                latest_por_documento[doc_id] = ds

        documentos_personales = [
            ds
            for ds in latest_por_documento.values()
            if not _es_categoria_archivo_final(ds.documento_requerido.categoria)
        ]

        documentos_salud = [
            ds
            for ds in documentos_personales
            if ds.documento_requerido_id in doc_salud_ids
        ]
        if doc_salud_ids and not documentos_salud:
            incidencias.append(
                {
                    "tipo": "faltante",
                    "detalle": "Falta el archivo de auto reporte de condiciones de salud.",
                }
            )

        for doc_subido in documentos_personales:
            if doc_subido.estado == "rechazado":
                incidencias.append(
                    {
                        "tipo": "rechazado",
                        "tipo_correccion": "documento",
                        "documento_subido_id": doc_subido.id,
                        "detalle": (
                            f"{doc_subido.documento_requerido.titulo}: "
                            f"{doc_subido.observaciones_revision or 'Documento mal diligenciado.'}"
                        ),
                    }
                )

        if getattr(
            asistente, "estado_autorizacion_padres", ""
        ) == "rechazado" and getattr(
            asistente, "observaciones_autorizacion_padres", ""
        ):
            incidencias.append(
                {
                    "tipo": "rechazado",
                    "tipo_correccion": "autorizacion_padres",
                    "detalle": (
                        "Autorización de padres: "
                        f"{asistente.observaciones_autorizacion_padres}"
                    ),
                }
            )

        if incidencias:
            asistentes_con_alertas.append(
                {
                    "id": asistente.id,
                    "nombre": asistente.nombre_completo,
                    "documento": f"{asistente.get_tipo_documento_display()} {asistente.numero_documento}",
                    "incidencias": incidencias,
                }
            )

    archivos_finales_estado = []
    for doc_final in docs_finales_requeridos:
        filtros = {"documento_requerido": doc_final}
        if tipo_visita == "interna":
            filtros["asistente_interna__visita"] = visita
        else:
            filtros["asistente_externa__visita"] = visita

        ultimo_archivo = (
            DocumentoSubidoAsistente.objects.filter(**filtros)
            .order_by("-fecha_subida")
            .first()
        )

        if not ultimo_archivo:
            estado = "faltante"
            detalle = "No se ha cargado este archivo final."
        elif ultimo_archivo.estado == "rechazado":
            estado = "rechazado"
            detalle = (
                ultimo_archivo.observaciones_revision or "Archivo final rechazado."
            )
        elif ultimo_archivo.estado == "pendiente":
            estado = "pendiente"
            detalle = "Archivo final cargado y en revisión."
        else:
            estado = "aprobado"
            detalle = "Archivo final aprobado."

        archivos_finales_estado.append(
            {
                "documento_requerido_id": doc_final.id,
                "titulo": doc_final.titulo,
                "categoria": doc_final.get_categoria_display(),
                "estado": estado,
                "detalle": detalle,
            }
        )

    archivos_finales_con_alerta = [
        a
        for a in archivos_finales_estado
        if a["estado"] == "rechazado"
        or (a["estado"] == "faltante" and mostrar_faltantes_finales_como_alerta)
    ]
    archivos_finales_rechazados = [
        a for a in archivos_finales_estado if a["estado"] == "rechazado"
    ]
    mostrar_estado_archivos_finales = any(
        a["estado"] != "aprobado" for a in archivos_finales_estado
    )
    total_incidencias_asistentes = sum(
        len(item["incidencias"]) for item in asistentes_con_alertas
    )

    return {
        "asistentes_con_alertas": asistentes_con_alertas,
        "archivos_finales_estado": archivos_finales_estado,
        "archivos_finales_con_alerta": archivos_finales_con_alerta,
        "archivos_finales_rechazados": archivos_finales_rechazados,
        "mostrar_estado_archivos_finales": mostrar_estado_archivos_finales,
        "total_alertas": total_incidencias_asistentes
        + len(archivos_finales_con_alerta),
        "hay_alertas": bool(asistentes_con_alertas or archivos_finales_con_alerta),
    }


def _solicitud_final_historica_interna(visita):
    """Determina si la visita ya fue enviada a revision al menos una vez."""
    if visita.estado in ["documentos_enviados", "en_revision_documentos", "confirmada"]:
        return True

    if HistorialAccionVisitaInterna.objects.filter(
        visita=visita,
        descripcion__icontains=MARCADOR_SOLICITUD_FINAL_ENVIADA,
    ).exists():
        return True

    if HistorialAccionVisitaInterna.objects.filter(
        visita=visita,
        tipo_accion__in=["inicio_revision", "devolucion_correccion", "confirmacion"],
    ).exists():
        return True

    if visita.asistentes.filter(
        estado__in=["documentos_aprobados", "documentos_rechazados"]
    ).exists():
        return True

    return DocumentoSubidoAsistente.objects.filter(
        asistente_interna__visita=visita,
        estado__in=["aprobado", "rechazado"],
    ).exists()




def get_sesion_instructor(request):
    """
    Lee los datos de sesión del panel_visitante.
    Retorna (correo, documento) o (None, None) si no está autenticado como interno.
    """
    if not request.session.get("responsable_autenticado"):
        return None, None
    if request.session.get("responsable_rol") != "interno":
        return None, None
    correo = request.session.get("responsable_correo")
    documento = request.session.get("responsable_documento")
    if not correo or not documento:
        return None, None
    return correo, documento


def instructor_interno_required(view_func):
    """
    Decorador: verifica sesión de panel_visitante con rol interno.
    """

    def wrapper(request, *args, **kwargs):
        correo, documento = get_sesion_instructor(request)
        if not correo:
            messages.warning(
                request,
                "Debes iniciar sesión como usuario interno para acceder a este panel.",
            )
            return redirect("panel_visitante:login_responsable")
        return view_func(request, *args, **kwargs)

    return wrapper


def _obtener_propietario_instructor(request):
    """Resuelve el usuario propietario para aislar datos por instructor interno."""
    correo = (request.session.get("responsable_correo") or "").strip().lower()
    documento = (request.session.get("responsable_documento") or "").strip()
    nombre = (request.session.get("responsable_nombre") or "").strip()
    apellido = (request.session.get("responsable_apellido") or "").strip()

    if not documento:
        return None

    UserModel = get_user_model()
    username = f"interno_{documento}"
    user, _ = UserModel.objects.get_or_create(
        username=username,
        defaults={
            "email": correo,
            "first_name": nombre,
            "last_name": apellido,
        },
    )

    update_fields = []
    if correo and user.email != correo:
        user.email = correo
        update_fields.append("email")
    if nombre and user.first_name != nombre:
        user.first_name = nombre
        update_fields.append("first_name")
    if apellido and user.last_name != apellido:
        user.last_name = apellido
        update_fields.append("last_name")
    if update_fields:
        user.save(update_fields=update_fields)

    return user




@instructor_interno_required
def panel_instructor_interno(request):
    correo, documento = get_sesion_instructor(request)
    owner_user = _obtener_propietario_instructor(request)
    visitas = VisitaInterna.objects.filter(correo_responsable__iexact=correo).order_by(
        "-fecha_solicitud"
    )
    context = {
        "correo": correo,
        "documento": documento,
        "nombre": request.session.get("responsable_nombre", ""),
        "apellido": request.session.get("responsable_apellido", ""),
        "visitas": visitas[:5],
        "total_programas": Programa.objects.filter(
            activo=True, creado_por=owner_user
        ).count(),
        "total_fichas": Ficha.objects.filter(
            activa=True, creado_por=owner_user
        ).count(),
        "total_visitas": visitas.count(),
    }
    return render(request, "panel_instructor_interno/panel.html", context)




@instructor_interno_required
def reservar_visita_interna(request):
    correo, documento = get_sesion_instructor(request)
    owner_user = _obtener_propietario_instructor(request)
    fichas = (
        Ficha.objects.filter(activa=True, creado_por=owner_user)
        .select_related("programa")
        .order_by("numero")
    )
    fichas_data = {str(f.numero): f.cantidad_aprendices for f in fichas}

    nombre = request.session.get("responsable_nombre", "")
    apellido = request.session.get("responsable_apellido", "")
    tipo_documento = request.session.get("responsable_tipo_documento", "CC")
    telefono = request.session.get("responsable_telefono", "")
    nombre_completo = f"{nombre} {apellido}".strip()

    if request.method == "POST":
        form = VisitaInternaInstructorForm(request.POST, owner_user=owner_user)
        if form.is_valid():
            numero_ficha = form.cleaned_data.get("numero_ficha")
            ficha = (
                Ficha.objects.select_related("programa")
                .filter(
                    numero=numero_ficha,
                    activa=True,
                    programa__activo=True,
                    creado_por=owner_user,
                )
                .first()
            )

            if not ficha:
                form.add_error(
                    "numero_ficha",
                    "La ficha seleccionada no es válida o está inactiva.",
                )
                context = {
                    "form": form,
                    "fichas": fichas,
                    "fichas_data": fichas_data,
                    "correo": correo,
                    "titulo": "Reservar Visita Interna",
                }
                return render(
                    request, "panel_instructor_interno/reservar_visita.html", context
                )

            visita = form.save(commit=False)
            visita.nombre_programa = ficha.programa.nombre
            visita.numero_ficha = ficha.numero

            if ficha.cantidad_aprendices < 1:
                form.add_error(
                    "cantidad_aprendices",
                    "La ficha seleccionada tiene cantidad de aprendices inválida. Actualízala en Fichas y Programas.",
                )
                context = {
                    "form": form,
                    "fichas": fichas,
                    "fichas_data": fichas_data,
                    "correo": correo,
                    "titulo": "Reservar Visita Interna",
                }
                return render(
                    request, "panel_instructor_interno/reservar_visita.html", context
                )

            visita.cantidad_aprendices = ficha.cantidad_aprendices
            visita.correo_responsable = correo
            visita.documento_responsable = documento
            visita.estado = "enviada_coordinacion"

            if not (visita.fecha_visita and visita.hora_inicio and visita.hora_fin):
                form.add_error(
                    None,
                    "Debes seleccionar fecha y horario antes de enviar la solicitud.",
                )
            elif visita.hora_inicio >= visita.hora_fin:
                form.add_error(None, "El horario seleccionado no es válido.")
            else:
                try:
                    with transaction.atomic():
                        visita.save()

                        if not ReservaHorario.horario_disponible(
                            visita.fecha_visita,
                            visita.hora_inicio,
                            visita.hora_fin,
                        ):
                            raise ValueError(
                                "El horario seleccionado ya no está disponible. Por favor elige otro horario."
                            )

                        reserva = ReservaHorario.crear_reserva_interna(visita)
                        if not reserva:
                            raise ValueError(
                                "No fue posible bloquear el horario seleccionado. Intenta nuevamente."
                            )
                except ValueError as exc:
                    form.add_error(None, str(exc))
                else:
                    try:
                        subject = "Confirmación: solicitud de visita interna enviada"
                        context = {
                            "responsable_nombre": nombre_completo,
                            "responsable": visita.responsable,
                            "documento_responsable": visita.documento_responsable,
                            "nombre_programa": visita.nombre_programa,
                            "numero_ficha": visita.numero_ficha,
                            "fecha_visita": (
                                visita.fecha_visita.strftime("%d/%m/%Y")
                                if visita.fecha_visita
                                else "Por definir"
                            ),
                            "hora_programada": (
                                (
                                    visita.hora_inicio.strftime("%I:%M %p")
                                    + " - "
                                    + visita.hora_fin.strftime("%I:%M %p")
                                )
                                if visita.hora_inicio and visita.hora_fin
                                else "Por definir"
                            ),
                        }
                        html_content = render_to_string(
                            "emails/solicitud_visita_interna.html", context
                        )
                        text_content = strip_tags(html_content)
                        msg = EmailMultiAlternatives(
                            subject,
                            text_content,
                            settings.DEFAULT_FROM_EMAIL,
                            [visita.correo_responsable],
                        )
                        msg.attach_alternative(html_content, "text/html")
                        msg.send(fail_silently=True)
                    except Exception:
                        pass

                    messages.success(
                        request,
                        "✅ Solicitud enviada. El horario quedó bloqueado y está pendiente de aprobación por coordinación.",
                    )
                    return redirect("panel_instructor_interno:panel")
    else:
        form = VisitaInternaInstructorForm(
            initial={
                "responsable": nombre_completo,
                "tipo_documento_responsable": tipo_documento,
                "documento_responsable": documento,
                "correo_responsable": correo,
                "telefono_responsable": telefono,
            },
            owner_user=owner_user,
        )
    context = {
        "form": form,
        "fichas": fichas,
        "fichas_data": fichas_data,
        "correo": correo,
        "titulo": "Reservar Visita Interna",
    }
    return render(request, "panel_instructor_interno/reservar_visita.html", context)


@instructor_interno_required
def detalle_visita_interna(request, pk):
    correo, _ = get_sesion_instructor(request)
    visita = get_object_or_404(VisitaInterna, pk=pk, correo_responsable__iexact=correo)
    reprogramacion_pendiente = (
        HistorialReprogramacion.objects.filter(visita_interna=visita, completada=False)
        .order_by("-fecha_solicitud")
        .first()
    )

    from documentos.models import Documento

    documentos_disponibles = Documento.objects.all().order_by(
        "categoria", "-fecha_subida"
    )
    documentos_por_categoria = {}
    for doc in documentos_disponibles:
        cat_display = doc.get_categoria_display()
        if cat_display not in documentos_por_categoria:
            documentos_por_categoria[cat_display] = []
        documentos_por_categoria[cat_display].append(doc)

    documentos_finales_requeridos = []
    categorias_finales_agregadas = set()
    for doc in documentos_disponibles:
        categoria_norm = _normalizar_categoria_texto(doc.categoria)
        clave_categoria = None
        etiqueta = doc.categoria

        if "ats" in categoria_norm:
            clave_categoria = "ats"
            etiqueta = "ATS"
        elif "induccion y reinduccion" in categoria_norm:
            clave_categoria = "induccion"
            etiqueta = "Formato Inducción y Reinducción"
        elif "charla de seguridad" in categoria_norm and (
            "calestenia" in categoria_norm or "calistenia" in categoria_norm
        ):
            clave_categoria = "charla"
            etiqueta = "Charla de Seguridad y Calistenia"

        if not clave_categoria or clave_categoria in categorias_finales_agregadas:
            continue

        doc.etiqueta_requerida = etiqueta
        documentos_finales_requeridos.append(doc)
        categorias_finales_agregadas.add(clave_categoria)

    reporte_documental = construir_reporte_documental_visita(visita, "interna")
    documentos_finales_rechazados_ids = [
        item.get("documento_requerido_id")
        for item in reporte_documental.get("archivos_finales_rechazados", [])
        if item.get("documento_requerido_id") is not None
    ]
    categorias_finales_rechazadas = [
        item.get("categoria")
        for item in reporte_documental.get("archivos_finales_rechazados", [])
        if item.get("categoria")
    ]
    hay_alertas_documentales = bool(reporte_documental.get("hay_alertas"))
    estados_finales = reporte_documental.get("archivos_finales_estado", [])
    archivos_finales_faltantes = sum(
        1 for item in estados_finales if item.get("estado") == "faltante"
    )
    archivos_finales_completos = archivos_finales_faltantes == 0
    mostrar_boton_corregir_archivos_finales = any(
        item.get("estado") == "rechazado" for item in estados_finales
    )

    enviar_final_habilitado = (
        visita.estado == "aprobada_inicial"
        and visita.asistentes.exists()
        and archivos_finales_completos
        and not hay_alertas_documentales
    )

    mostrar_boton_subir_archivos_finales = (
        visita.estado == "aprobada_inicial"
        and visita.asistentes.exists()
        and not archivos_finales_completos
    )

    mostrar_reporte_diligenciamiento = (
        visita.estado in ["documentos_enviados", "en_revision_documentos", "confirmada"]
        or DocumentoSubidoAsistente.objects.filter(
            asistente_interna__visita=visita,
            documento_requerido__categoria__in=CATEGORIAS_ARCHIVOS_FINALES,
        ).exists()
    )

    solicitud_final_historica = _solicitud_final_historica_interna(visita)
    permite_editar_asistentes = (
        visita.estado == "aprobada_inicial" and not solicitud_final_historica
    )

    return render(
        request,
        "panel_instructor_interno/detalle_visita.html",
        {
            "visita": visita,
            "correo": correo,
            "documentos_por_categoria": documentos_por_categoria,
            "documentos_finales_requeridos": documentos_finales_requeridos,
            "documentos_finales_rechazados_ids": documentos_finales_rechazados_ids,
            "categorias_finales_rechazadas": categorias_finales_rechazadas,
            "reporte_documental": reporte_documental,
            "mostrar_reporte_diligenciamiento": mostrar_reporte_diligenciamiento,
            "reprogramacion_pendiente": reprogramacion_pendiente,
            "enviar_final_habilitado": enviar_final_habilitado,
            "mostrar_boton_subir_archivos_finales": mostrar_boton_subir_archivos_finales,
            "mostrar_boton_corregir_archivos_finales": mostrar_boton_corregir_archivos_finales,
            "solicitud_final_historica": solicitud_final_historica,
            "permite_editar_asistentes": permite_editar_asistentes,
        },
    )




@instructor_interno_required
def gestionar_programas(request):
    return redirect("panel_instructor_interno:gestionar_fichas")


@instructor_interno_required
def crear_programa(request):
    correo, _ = get_sesion_instructor(request)
    owner_user = _obtener_propietario_instructor(request)
    is_ajax = (
        request.GET.get("ajax") == "true"
        or request.headers.get("X-Requested-With") == "XMLHttpRequest"
    )

    if request.method == "POST":
        form = ProgramaForm(request.POST)
        if form.is_valid():
            programa = form.save(commit=False)
            programa.creado_por = owner_user
            programa.save()
            if is_ajax:
                return JsonResponse(
                    {
                        "success": True,
                        "message": f'✅ Programa "{programa.nombre}" creado correctamente.',
                    }
                )
            else:
                messages.success(
                    request, f'✅ Programa "{programa.nombre}" creado correctamente.'
                )
                return redirect("panel_instructor_interno:gestionar_fichas")
        else:
            if is_ajax:
                html = render(
                    request,
                    "panel_instructor_interno/form_programa_modal.html",
                    {
                        "form": form,
                        "titulo": "Crear Programa",
                        "accion": "Crear",
                        "correo": correo,
                    },
                ).content.decode("utf-8")
                return JsonResponse({"success": False, "html": html})
            return render(
                request,
                "panel_instructor_interno/form_programa.html",
                {
                    "form": form,
                    "titulo": "Crear Programa",
                    "accion": "Crear",
                    "correo": correo,
                },
            )
    else:
        form = ProgramaForm()

    if is_ajax:
        html = render(
            request,
            "panel_instructor_interno/form_programa_modal.html",
            {
                "form": form,
                "titulo": "Crear Programa",
                "accion": "Crear",
                "correo": correo,
            },
        ).content.decode("utf-8")
        return JsonResponse({"html": html})
    else:
        return render(
            request,
            "panel_instructor_interno/form_programa.html",
            {
                "form": form,
                "titulo": "Crear Programa",
                "accion": "Crear",
                "correo": correo,
            },
        )


@instructor_interno_required
def editar_programa(request, pk):
    correo, _ = get_sesion_instructor(request)
    owner_user = _obtener_propietario_instructor(request)
    programa = get_object_or_404(Programa, pk=pk, creado_por=owner_user)
    is_ajax = (
        request.GET.get("ajax") == "true"
        or request.headers.get("X-Requested-With") == "XMLHttpRequest"
    )

    if request.method == "POST":
        form = ProgramaForm(request.POST, instance=programa)
        if form.is_valid():
            form.save()
            if is_ajax:
                return JsonResponse(
                    {
                        "success": True,
                        "message": f'✅ Programa "{programa.nombre}" actualizado.',
                    }
                )
            else:
                messages.success(
                    request, f'✅ Programa "{programa.nombre}" actualizado.'
                )
                return redirect("panel_instructor_interno:gestionar_fichas")
        else:
            if is_ajax:
                html = render(
                    request,
                    "panel_instructor_interno/form_programa_modal.html",
                    {
                        "form": form,
                        "titulo": "Editar Programa",
                        "accion": "Actualizar",
                        "correo": correo,
                    },
                ).content.decode("utf-8")
                return JsonResponse({"success": False, "html": html})
            return render(
                request,
                "panel_instructor_interno/form_programa.html",
                {
                    "form": form,
                    "programa": programa,
                    "titulo": "Editar Programa",
                    "accion": "Actualizar",
                    "correo": correo,
                },
            )
    else:
        form = ProgramaForm(instance=programa)

    if is_ajax:
        html = render(
            request,
            "panel_instructor_interno/form_programa_modal.html",
            {
                "form": form,
                "titulo": "Editar Programa",
                "accion": "Actualizar",
                "correo": correo,
            },
        ).content.decode("utf-8")
        return JsonResponse({"html": html})
    else:
        return render(
            request,
            "panel_instructor_interno/form_programa.html",
            {
                "form": form,
                "programa": programa,
                "titulo": "Editar Programa",
                "accion": "Actualizar",
                "correo": correo,
            },
        )


@instructor_interno_required
def eliminar_programa(request, pk):
    correo, _ = get_sesion_instructor(request)
    owner_user = _obtener_propietario_instructor(request)
    programa = get_object_or_404(Programa, pk=pk, creado_por=owner_user)
    is_ajax = (
        request.GET.get("ajax") == "true"
        or request.headers.get("X-Requested-With") == "XMLHttpRequest"
    )

    if request.method == "POST":
        nombre = programa.nombre
        try:
            programa.delete()
            if is_ajax:
                return JsonResponse(
                    {"success": True, "message": f'🗑️ Programa "{nombre}" eliminado.'}
                )
            else:
                messages.success(request, f'🗑️ Programa "{nombre}" eliminado.')
        except Exception:
            if is_ajax:
                return JsonResponse(
                    {
                        "success": False,
                        "message": f'❌ No se puede eliminar "{nombre}" porque tiene fichas asociadas.',
                    }
                )
            else:
                messages.error(
                    request,
                    f'❌ No se puede eliminar "{nombre}" porque tiene fichas asociadas.',
                )
        return redirect("panel_instructor_interno:gestionar_fichas")

    if is_ajax:
        return JsonResponse(
            {
                "nombre": programa.nombre,
                "descripcion": programa.descripcion or "Sin descripción",
            }
        )
    else:
        return render(
            request,
            "panel_instructor_interno/confirmar_eliminar_programa.html",
            {
                "programa": programa,
                "correo": correo,
            },
        )




@instructor_interno_required
def gestionar_fichas(request):
    correo, _ = get_sesion_instructor(request)
    owner_user = _obtener_propietario_instructor(request)
    fichas = Ficha.objects.select_related("programa").filter(creado_por=owner_user)
    buscar = request.GET.get("buscar", "")
    if buscar:
        fichas = fichas.filter(
            Q(numero__icontains=buscar) | Q(programa__nombre__icontains=buscar)
        )
    return render(
        request,
        "panel_instructor_interno/gestionar_fichas.html",
        {
            "fichas": fichas,
            "buscar": buscar,
            "total": fichas.count(),
            "correo": correo,
        },
    )


@instructor_interno_required
def crear_ficha(request):
    correo, _ = get_sesion_instructor(request)
    owner_user = _obtener_propietario_instructor(request)
    is_ajax = (
        request.GET.get("ajax") == "true"
        or request.headers.get("X-Requested-With") == "XMLHttpRequest"
    )

    if request.method == "POST":
        form = FichaForm(request.POST, owner_user=owner_user)
        if form.is_valid():
            ficha = form.save(commit=False)
            ficha.creado_por = owner_user
            ficha.save()
            if is_ajax:
                return JsonResponse(
                    {
                        "success": True,
                        "message": f"✅ Registro de ficha y programa ({ficha.numero}) creado correctamente.",
                    }
                )
            else:
                messages.success(
                    request,
                    f"✅ Registro de ficha y programa ({ficha.numero}) creado correctamente.",
                )
                return redirect("panel_instructor_interno:gestionar_fichas")
        else:
            if is_ajax:
                html = render(
                    request,
                    "panel_instructor_interno/form_ficha_modal.html",
                    {
                        "form": form,
                        "titulo": "Crear Ficha y Programa",
                        "accion": "Crear",
                        "correo": correo,
                    },
                ).content.decode("utf-8")
                return JsonResponse({"success": False, "html": html})
            return render(
                request,
                "panel_instructor_interno/form_ficha.html",
                {
                    "form": form,
                    "titulo": "Crear Ficha y Programa",
                    "accion": "Crear",
                    "correo": correo,
                },
            )
    else:
        form = FichaForm(owner_user=owner_user)

    if is_ajax:
        html = render(
            request,
            "panel_instructor_interno/form_ficha_modal.html",
            {
                "form": form,
                "titulo": "Crear Ficha y Programa",
                "accion": "Crear",
                "correo": correo,
            },
        ).content.decode("utf-8")
        return JsonResponse({"html": html})
    else:
        return render(
            request,
            "panel_instructor_interno/form_ficha.html",
            {
                "form": form,
                "titulo": "Crear Ficha y Programa",
                "accion": "Crear",
                "correo": correo,
            },
        )


@instructor_interno_required
def editar_ficha(request, pk):
    correo, _ = get_sesion_instructor(request)
    owner_user = _obtener_propietario_instructor(request)
    ficha = get_object_or_404(Ficha, pk=pk, creado_por=owner_user)
    is_ajax = (
        request.GET.get("ajax") == "true"
        or request.headers.get("X-Requested-With") == "XMLHttpRequest"
    )

    if request.method == "POST":
        form = FichaForm(request.POST, instance=ficha, owner_user=owner_user)
        if form.is_valid():
            form.save()
            if is_ajax:
                return JsonResponse(
                    {
                        "success": True,
                        "message": f"✅ Registro de ficha y programa ({ficha.numero}) actualizado.",
                    }
                )
            else:
                messages.success(
                    request,
                    f"✅ Registro de ficha y programa ({ficha.numero}) actualizado.",
                )
                return redirect("panel_instructor_interno:gestionar_fichas")
        else:
            if is_ajax:
                html = render(
                    request,
                    "panel_instructor_interno/form_ficha_modal.html",
                    {
                        "form": form,
                        "titulo": "Editar Ficha y Programa",
                        "accion": "Actualizar",
                        "correo": correo,
                    },
                ).content.decode("utf-8")
                return JsonResponse({"success": False, "html": html})
            return render(
                request,
                "panel_instructor_interno/form_ficha.html",
                {
                    "form": form,
                    "ficha": ficha,
                    "titulo": "Editar Ficha y Programa",
                    "accion": "Actualizar",
                    "correo": correo,
                },
            )
    else:
        form = FichaForm(instance=ficha, owner_user=owner_user)

    if is_ajax:
        html = render(
            request,
            "panel_instructor_interno/form_ficha_modal.html",
            {
                "form": form,
                "titulo": "Editar Ficha y Programa",
                "accion": "Actualizar",
                "correo": correo,
            },
        ).content.decode("utf-8")
        return JsonResponse({"html": html})
    else:
        return render(
            request,
            "panel_instructor_interno/form_ficha.html",
            {
                "form": form,
                "ficha": ficha,
                "titulo": "Editar Ficha y Programa",
                "accion": "Actualizar",
                "correo": correo,
            },
        )


@instructor_interno_required
def eliminar_ficha(request, pk):
    correo, _ = get_sesion_instructor(request)
    owner_user = _obtener_propietario_instructor(request)
    ficha = get_object_or_404(Ficha, pk=pk, creado_por=owner_user)
    is_ajax = (
        request.GET.get("ajax") == "true"
        or request.headers.get("X-Requested-With") == "XMLHttpRequest"
    )

    if request.method == "POST":
        numero = ficha.numero
        try:
            ficha.delete()
            if is_ajax:
                return JsonResponse(
                    {"success": True, "message": f"🗑️ Ficha {numero} eliminada."}
                )
            else:
                messages.success(request, f"🗑️ Ficha {numero} eliminada.")
        except Exception as e:
            if is_ajax:
                return JsonResponse(
                    {
                        "success": False,
                        "message": f"❌ No se puede eliminar la ficha. Error: {str(e)}",
                    }
                )
            else:
                messages.error(request, f"❌ No se puede eliminar la ficha.")
        return redirect("panel_instructor_interno:gestionar_fichas")

    if is_ajax:
        return JsonResponse({"numero": ficha.numero, "programa": ficha.programa.nombre})
    else:
        return render(
            request,
            "panel_instructor_interno/confirmar_eliminar_ficha.html",
            {
                "ficha": ficha,
                "correo": correo,
            },
        )




@instructor_interno_required
def listar_fichas_aprendices(request):
    """
    Ruta legacy. Redirige al módulo principal de fichas.
    """
    return redirect("panel_instructor_interno:gestionar_fichas")


from django.views.decorators.clickjacking import xframe_options_exempt


@xframe_options_exempt
@instructor_interno_required
def ver_documento_aprendiz_inline(request, aprendiz_id, campo):
    """Sirve un campo de archivo de un Aprendiz (documento_identidad o documento_adicional) inline."""
    import mimetypes
    from django.http import FileResponse, HttpResponseNotFound, HttpResponseBadRequest

    campos_permitidos = ["documento_identidad", "documento_adicional"]
    if campo not in campos_permitidos:
        return HttpResponseBadRequest("Campo no válido")

    owner_user = _obtener_propietario_instructor(request)
    aprendiz = get_object_or_404(Aprendiz, id=aprendiz_id, ficha__creado_por=owner_user)
    archivo = getattr(aprendiz, campo)
    if not archivo or not archivo.storage.exists(archivo.name):
        return HttpResponseNotFound("El archivo no existe")

    nombre = os.path.basename(archivo.name)
    f = archivo.open("rb")
    tipo_mime, _ = mimetypes.guess_type(archivo.name)
    if not tipo_mime:
        tipo_mime = "application/octet-stream"

    response = FileResponse(f, content_type=tipo_mime)
    response["Content-Disposition"] = f'inline; filename="{nombre}"'
    response["X-Frame-Options"] = "SAMEORIGIN"
    return response


@instructor_interno_required
def detalle_aprendices_ficha(request, pk):
    """
    Muestra todos los aprendices asociados a una ficha.
    Permite crear, editar y eliminar aprendices.
    """
    correo, _ = get_sesion_instructor(request)
    owner_user = _obtener_propietario_instructor(request)
    ficha = get_object_or_404(Ficha, pk=pk, creado_por=owner_user)

    if request.method == "POST" and request.POST.get("accion") == "carga_masiva_aprendices":
        archivo = request.FILES.get("archivo_carga_masiva")
        if not archivo:
            messages.error(request, "Seleccione un archivo para realizar la carga masiva.")
            return redirect("panel_instructor_interno:detalle_aprendices_ficha", pk=ficha.id)

        filas, errores_lectura = _extraer_filas_importacion_aprendices(archivo)
        if errores_lectura:
            for error in errores_lectura:
                messages.error(request, error)
            return redirect("panel_instructor_interno:detalle_aprendices_ficha", pk=ficha.id)

        resumen = _procesar_carga_masiva_aprendices(ficha, filas)
        if resumen["creados"]:
            messages.success(
                request,
                f"Carga masiva completada. Se crearon {resumen['creados']} aprendices.",
            )

        if resumen["duplicados"]:
            messages.warning(
                request,
                f"Se omitieron {resumen['duplicados']} filas por documento repetido en la ficha.",
            )

        if resumen.get("omitidos_cupo"):
            messages.warning(
                request,
                (
                    f"Se omitieron {resumen['omitidos_cupo']} filas porque la ficha "
                    f"alcanzó su límite de {ficha.cantidad_aprendices} aprendices."
                ),
            )

        for error in resumen["errores"][:10]:
            messages.error(request, error)

        if len(resumen["errores"]) > 10:
            messages.error(
                request,
                f"Se ocultaron {len(resumen['errores']) - 10} errores adicionales para mantener legible el resultado.",
            )

        return redirect("panel_instructor_interno:detalle_aprendices_ficha", pk=ficha.id)

    if request.method == "POST" and request.POST.get("accion") == "subir_reporte_salud_pendiente":
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        try:
            aprendiz_id = request.POST.get("aprendiz_id", "").strip()
            archivo_salud = request.FILES.get("documento_salud")
            archivo_autorizacion = request.FILES.get("documento_autorizacion_padres")

            if not aprendiz_id:
                mensaje = "ID de aprendiz inválido."
                if is_ajax:
                    return JsonResponse({"success": False, "message": mensaje}, status=400)
                messages.error(request, mensaje)
                return redirect("panel_instructor_interno:detalle_aprendices_ficha", pk=ficha.id)

            aprendiz = get_object_or_404(Aprendiz, pk=aprendiz_id, ficha=ficha)

            if not archivo_salud and not archivo_autorizacion:
                mensaje = "Seleccione al menos un archivo para continuar."
                if is_ajax:
                    return JsonResponse({"success": False, "message": mensaje}, status=400)
                messages.error(request, mensaje)
                return redirect("panel_instructor_interno:detalle_aprendices_ficha", pk=ficha.id)

            falta_salud = not (bool(aprendiz.documento_adicional) or aprendiz.documentos_subidos.filter(
                documento_requerido__categoria=CATEGORIA_DOC_SALUD
            ).exists())
            
            falta_autorizacion = False
            if aprendiz.tipo_documento == "TI":
                falta_autorizacion = not any(
                    "autorizacion padres" in _normalizar_categoria_texto(doc.documento_requerido.categoria)
                    for doc in aprendiz.documentos_subidos.all()
                )

            if aprendiz.tipo_documento == "TI" and falta_salud and falta_autorizacion:
                if not archivo_salud or not archivo_autorizacion:
                    mensaje = "Aprendiz con Tarjeta de Identidad: Debe subir AMBOS archivos (reporte de salud y autorización de padres)."
                    if is_ajax:
                        return JsonResponse({"success": False, "message": mensaje}, status=400)
                    messages.error(request, mensaje)
                    return redirect("panel_instructor_interno:detalle_aprendices_ficha", pk=ficha.id)

            def _validar_archivo(archivo):
                extension = os.path.splitext(archivo.name)[1].lower()
                if extension not in EXTENSIONES_DOCUMENTOS_APRENDIZ:
                    return "Formato no permitido. Solo se admiten archivos PDF o Word (.doc, .docx)."
                if getattr(archivo, "size", 0) > MAX_TAMANO_DOCUMENTO_APRENDIZ:
                    return "El archivo supera el tamaño máximo permitido de 10MB."
                return None

            if archivo_salud:
                error_archivo = _validar_archivo(archivo_salud)
                if error_archivo:
                    if is_ajax:
                        return JsonResponse({"success": False, "message": error_archivo}, status=400)
                    messages.error(request, error_archivo)
                    return redirect("panel_instructor_interno:detalle_aprendices_ficha", pk=ficha.id)

            if archivo_autorizacion:
                error_archivo = _validar_archivo(archivo_autorizacion)
                if error_archivo:
                    if is_ajax:
                        return JsonResponse({"success": False, "message": error_archivo}, status=400)
                    messages.error(request, error_archivo)
                    return redirect("panel_instructor_interno:detalle_aprendices_ficha", pk=ficha.id)

            cambios_realizados = []

            if archivo_salud:
                aprendiz.documento_adicional = archivo_salud
                aprendiz.save(update_fields=["documento_adicional"])
                cambios_realizados.append("reporte de salud")

            if archivo_autorizacion:
                doc_autorizacion = Documento.objects.filter(
                    categoria=CATEGORIA_AUTORIZACION_PADRES
                ).order_by("-fecha_subida").first()

                if not doc_autorizacion:
                    docs_autorizacion = [
                        doc
                        for doc in Documento.objects.all().order_by("-fecha_subida")
                        if "autorizacion padres" in _normalizar_categoria_texto(doc.categoria)
                    ]
                    doc_autorizacion = docs_autorizacion[0] if docs_autorizacion else None

                if not doc_autorizacion:
                    mensaje = (
                        "No se encontró un documento configurado para 'Formato Autorización Padres de Familia'."
                    )
                    if is_ajax:
                        return JsonResponse({"success": False, "message": mensaje}, status=400)
                    messages.error(request, mensaje)
                    return redirect("panel_instructor_interno:detalle_aprendices_ficha", pk=ficha.id)

                DocumentoSubidoAprendiz.objects.update_or_create(
                    aprendiz=aprendiz,
                    documento_requerido=doc_autorizacion,
                    defaults={
                        "archivo": archivo_autorizacion,
                        "estado": "pendiente",
                        "observaciones_revision": "",
                    },
                )
                cambios_realizados.append("autorización de padres")

            faltantes_actualizados = []
            tiene_doc_salud = bool(aprendiz.documento_adicional) or aprendiz.documentos_subidos.filter(
                documento_requerido__categoria=CATEGORIA_DOC_SALUD
            ).exists()
            if not tiene_doc_salud:
                faltantes_actualizados.append("Reporte de condiciones de salud")

            if aprendiz.tipo_documento == "TI":
                tiene_autorizacion_padres = any(
                    "autorizacion padres"
                    in _normalizar_categoria_texto(doc.documento_requerido.categoria)
                    for doc in aprendiz.documentos_subidos.select_related("documento_requerido")
                )
                if not tiene_autorizacion_padres:
                    faltantes_actualizados.append("Autorización de padres")

            mensaje = (
                f'Documento(s) cargado(s) correctamente para "{aprendiz.get_nombre_completo()}": '
                + ", ".join(cambios_realizados)
                + "."
            )
            if is_ajax:
                return JsonResponse(
                    {
                        "success": True,
                        "message": mensaje,
                        "aprendiz_id": aprendiz.id,
                        "faltantes": faltantes_actualizados,
                    }
                )
            
            messages.success(request, mensaje)
            return redirect("panel_instructor_interno:detalle_aprendices_ficha", pk=ficha.id)
        
        except Aprendiz.DoesNotExist:
            mensaje = "Aprendiz no encontrado."
            if is_ajax:
                return JsonResponse({"success": False, "message": mensaje}, status=404)
            messages.error(request, mensaje)
            return redirect("panel_instructor_interno:detalle_aprendices_ficha", pk=ficha.id)
        
        except Exception as error:
            mensaje = f"Error al guardar el archivo: {str(error)}"
            if is_ajax:
                return JsonResponse({"success": False, "message": mensaje}, status=500)
            messages.error(request, mensaje)
            return redirect("panel_instructor_interno:detalle_aprendices_ficha", pk=ficha.id)

    aprendices = ficha.aprendices.all().order_by("apellido", "nombre")
    estado_filtro = request.GET.get("estado", "")

    if estado_filtro:
        aprendices = aprendices.filter(estado=estado_filtro)

    stats = {
        "total": ficha.aprendices.count(),
        "activos": ficha.aprendices.filter(estado="activo").count(),
        "inactivos": ficha.aprendices.filter(estado="inactivo").count(),
        "retirados": ficha.aprendices.filter(estado="retirado").count(),
    }

    aprendices_faltantes_docs = []
    aprendices_para_revision = ficha.aprendices.prefetch_related(
        "documentos_subidos__documento_requerido"
    )
    for aprendiz in aprendices_para_revision:
        faltantes = []

        tiene_doc_salud = bool(aprendiz.documento_adicional) or aprendiz.documentos_subidos.filter(
            documento_requerido__categoria=CATEGORIA_DOC_SALUD
        ).exists()
        if not tiene_doc_salud:
            faltantes.append("Reporte de condiciones de salud")

        tiene_autorizacion_padres = False
        if aprendiz.tipo_documento == "TI":
            tiene_autorizacion_padres = any(
                "autorizacion padres"
                in _normalizar_categoria_texto(doc.documento_requerido.categoria)
                for doc in aprendiz.documentos_subidos.all()
            )
            if not tiene_autorizacion_padres:
                faltantes.append("Autorización de padres")

        if faltantes:
            aprendices_faltantes_docs.append(
                {
                    "id": aprendiz.id,
                    "nombre": aprendiz.get_nombre_completo(),
                    "tipo_documento": aprendiz.tipo_documento,
                    "documento": f"{aprendiz.get_tipo_documento_display()} {aprendiz.numero_documento}",
                    "faltantes": faltantes,
                    "falta_salud": not tiene_doc_salud,
                    "falta_autorizacion_padres": (
                        aprendiz.tipo_documento == "TI" and not tiene_autorizacion_padres
                    ),
                }
            )

    context = {
        "ficha": ficha,
        "aprendices": aprendices,
        "stats": stats,
        "estado_filtro": estado_filtro,
        "correo": correo,
        "aprendices_faltantes_docs": aprendices_faltantes_docs,
        "total_faltantes_docs": len(aprendices_faltantes_docs),
    }

    return render(
        request, "panel_instructor_interno/detalle_aprendices_ficha.html", context
    )


@instructor_interno_required
def descargar_plantilla_carga_masiva_aprendices(request, formato):
    formato = str(formato or "").lower()

    if formato == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = (
            'attachment; filename="plantilla_carga_masiva_aprendices.csv"'
        )
        writer = csv.DictWriter(response, fieldnames=COLUMNAS_CARGA_MASIVA_APRENDICES)
        writer.writeheader()
        writer.writerows(EJEMPLOS_CARGA_MASIVA_APRENDICES)
        return response

    if formato == "xlsx":
        try:
            from openpyxl import Workbook
        except Exception:
            messages.error(
                request,
                "No fue posible generar la plantilla Excel. Instale openpyxl.",
            )
            return redirect("panel_instructor_interno:gestionar_fichas")

        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla"
        ws.append(COLUMNAS_CARGA_MASIVA_APRENDICES)
        for ejemplo in EJEMPLOS_CARGA_MASIVA_APRENDICES:
            ws.append([ejemplo.get(col, "") for col in COLUMNAS_CARGA_MASIVA_APRENDICES])

        salida = io.BytesIO()
        wb.save(salida)
        salida.seek(0)

        response = HttpResponse(
            salida.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            'attachment; filename="plantilla_carga_masiva_aprendices.xlsx"'
        )
        return response

    if formato == "pdf":
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas

        output = io.BytesIO()
        pdf = canvas.Canvas(output, pagesize=letter)

        y = 760
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(40, y, "Plantilla carga masiva de aprendices")
        y -= 22
        pdf.setFont("Helvetica", 9)
        pdf.drawString(
            40,
            y,
            "Cada fila debe conservar este orden y usar ';' como separador:",
        )
        y -= 14
        pdf.drawString(40, y, "; ".join(COLUMNAS_CARGA_MASIVA_APRENDICES))
        y -= 22
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(40, y, "Ejemplos:")
        y -= 16
        pdf.setFont("Helvetica", 9)

        for ejemplo in EJEMPLOS_CARGA_MASIVA_APRENDICES:
            linea = ";".join(
                [str(ejemplo.get(col, "")) for col in COLUMNAS_CARGA_MASIVA_APRENDICES]
            )
            pdf.drawString(40, y, linea)
            y -= 14

        y -= 6
        pdf.drawString(
            40,
            y,
            "Nota: si usas PDF para importar, mantén una fila por línea con ese formato.",
        )

        pdf.showPage()
        pdf.save()
        output.seek(0)

        response = HttpResponse(output.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = (
            'attachment; filename="plantilla_carga_masiva_aprendices.pdf"'
        )
        return response

    messages.error(request, "Formato de plantilla no válido.")
    return redirect("panel_instructor_interno:gestionar_fichas")


@instructor_interno_required
def crear_aprendiz(request, ficha_id):
    """
    Crea un nuevo aprendiz para una ficha.
    """
    correo, _ = get_sesion_instructor(request)
    owner_user = _obtener_propietario_instructor(request)
    ficha = get_object_or_404(Ficha, pk=ficha_id, creado_por=owner_user)

    documentos_por_categoria = obtener_documentos_por_categoria()

    if request.method == "POST":
        form = AprendizForm(request.POST, request.FILES, ficha=ficha)
        errores_archivos = validar_archivos_documentos_aprendiz(request.FILES)

        if form.is_valid() and not errores_archivos:
            aprendiz = form.save(commit=False)
            aprendiz.ficha = ficha
            try:
                with transaction.atomic():
                    aprendiz.save()

                    for categoria, docs in documentos_por_categoria.items():
                        if (
                            categoria
                            == "👩🏻‍⚕️ Formato Auto Reporte Condiciones de Salud"
                        ):
                            for doc in docs:
                                archivo = request.FILES.get(f"documento_{doc.id}")
                                if archivo:
                                    DocumentoSubidoAprendiz.objects.create(
                                        documento_requerido=doc,
                                        aprendiz=aprendiz,
                                        archivo=archivo,
                                        estado="pendiente",
                                    )

                messages.success(
                    request,
                    f'✅ Aprendiz "{aprendiz.get_nombre_completo()}" registrado correctamente con documentos.',
                )
                return redirect(
                    "panel_instructor_interno:detalle_aprendices_ficha", pk=ficha.id
                )
            except IntegrityError:
                messages.error(
                    request,
                    f"❌ Ya existe un aprendiz con el documento {aprendiz.numero_documento} en esta ficha. "
                    f"Cada documento debe ser único por ficha.",
                )
        else:
            for error in errores_archivos:
                messages.error(request, error)
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"❌ {field}: {error}")
    else:
        form = AprendizForm(ficha=ficha)

    context = {
        "form": form,
        "ficha": ficha,
        "correo": correo,
        "titulo": "Registrar Aprendiz",
        "documentos_por_categoria": documentos_por_categoria,
    }

    return render(request, "panel_instructor_interno/form_aprendiz.html", context)


@instructor_interno_required
def editar_aprendiz(request, pk):
    """
    Edita un aprendiz existente y sus documentos.
    """
    correo, _ = get_sesion_instructor(request)
    owner_user = _obtener_propietario_instructor(request)
    aprendiz = get_object_or_404(Aprendiz, pk=pk, ficha__creado_por=owner_user)
    ficha = aprendiz.ficha

    documentos_por_categoria = obtener_documentos_por_categoria()

    if request.method == "POST":
        form = AprendizForm(request.POST, request.FILES, instance=aprendiz, ficha=ficha)
        errores_archivos = validar_archivos_documentos_aprendiz(request.FILES)

        if form.is_valid() and not errores_archivos:
            try:
                with transaction.atomic():
                    form.save()

                    for categoria, docs in documentos_por_categoria.items():
                        if (
                            categoria
                            == "👩🏻‍⚕️ Formato Auto Reporte Condiciones de Salud"
                        ):
                            for doc in docs:
                                archivo = request.FILES.get(f"documento_{doc.id}")
                                if archivo:
                                    DocumentoSubidoAprendiz.objects.filter(
                                        documento_requerido=doc, aprendiz=aprendiz
                                    ).delete()
                                    DocumentoSubidoAprendiz.objects.create(
                                        documento_requerido=doc,
                                        aprendiz=aprendiz,
                                        archivo=archivo,
                                        estado="pendiente",
                                    )

                messages.success(
                    request,
                    f'✅ Aprendiz "{aprendiz.get_nombre_completo()}" actualizado.',
                )
                return redirect(
                    "panel_instructor_interno:detalle_aprendices_ficha", pk=ficha.id
                )
            except IntegrityError:
                messages.error(
                    request,
                    f"❌ Ya existe un aprendiz con el documento {aprendiz.numero_documento} en esta ficha. "
                    f"Cada documento debe ser único por ficha.",
                )
        else:
            for error in errores_archivos:
                messages.error(request, error)
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"❌ {field}: {error}")
    else:
        form = AprendizForm(instance=aprendiz, ficha=ficha)

    context = {
        "form": form,
        "aprendiz": aprendiz,
        "ficha": ficha,
        "correo": correo,
        "titulo": "Editar Aprendiz",
        "documentos_por_categoria": documentos_por_categoria,
    }

    return render(request, "panel_instructor_interno/form_aprendiz.html", context)


@instructor_interno_required
def eliminar_aprendiz(request, pk):
    """
    Elimina un aprendiz.
    Si el aprendiz fue registrado en alguna visita, también se elimina de la visita.
    Soporta GET para obtener los datos para confirmar en modal, y POST para ejecutar eliminación.
    """
    correo, _ = get_sesion_instructor(request)
    owner_user = _obtener_propietario_instructor(request)
    aprendiz = get_object_or_404(Aprendiz, pk=pk, ficha__creado_por=owner_user)
    ficha = aprendiz.ficha
    es_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"

    if request.method == "POST":
        nombre_completo = aprendiz.get_nombre_completo()
        numero_documento = aprendiz.numero_documento
        asistentes_eliminados = 0
        
        try:
            from visitaInterna.models import VisitaInterna, AsistenteVisitaInterna
            
            visitas_ficha = VisitaInterna.objects.filter(
                numero_ficha=ficha.numero,
                correo_responsable__iexact=correo
            )
            
            for visita in visitas_ficha:
                asistentes = AsistenteVisitaInterna.objects.filter(
                    visita=visita,
                    numero_documento=numero_documento
                )
                asistentes_eliminados += asistentes.count()
                asistentes.delete()
        except Exception as e:
            if es_ajax:
                return JsonResponse(
                    {
                        "success": False,
                        "error": f"⚠️ Hubo un problema al sincronizar las visitas: {str(e)}"
                    },
                    status=400
                )
            messages.warning(
                request,
                f"⚠️ Aprendiz eliminado pero hubo un problema al sincronizar las visitas: {str(e)}"
            )
        
        aprendiz.delete()
        
        mensaje = (
            f'🗑️ Aprendiz "{nombre_completo}" eliminado de la ficha'
            + (f' y de {asistentes_eliminados} visita(s).' if asistentes_eliminados > 0 else '.')
        )
        
        if es_ajax:
            return JsonResponse(
                {
                    "success": True,
                    "message": mensaje,
                    "ficha_id": ficha.id
                }
            )
        
        messages.success(request, mensaje)
        return redirect(
            "panel_instructor_interno:detalle_aprendices_ficha", pk=ficha.id
        )

    if es_ajax:
        return JsonResponse(
            {
                "id": aprendiz.id,
                "nombre_completo": aprendiz.get_nombre_completo(),
                "numero_documento": f"{aprendiz.get_tipo_documento_display()} {aprendiz.numero_documento}",
            }
        )

    context = {
        "aprendiz": aprendiz,
        "ficha": ficha,
        "correo": correo,
    }

    return render(
        request, "panel_instructor_interno/confirmar_eliminar_aprendiz.html", context
    )




@instructor_interno_required
def obtener_aprendices_ficha_json(request, ficha_id):
    """
    Retorna lista JSON de aprendices de una ficha.
    Usado para prellenar datos en formulario de visitas.
    """
    try:
        owner_user = _obtener_propietario_instructor(request)
        ficha = get_object_or_404(Ficha, pk=ficha_id, creado_por=owner_user)
        aprendices = ficha.aprendices.filter(estado="activo").values(
            "id",
            "nombre",
            "apellido",
            "tipo_documento",
            "numero_documento",
            "correo",
            "telefono",
        )
        return JsonResponse(
            {
                "success": True,
                "aprendices": list(aprendices),
                "total": aprendices.count(),
            }
        )
    except Exception as e:
        return JsonResponse(
            {
                "success": False,
                "error": str(e),
            },
            status=400,
        )


@instructor_interno_required
def registrar_aprendices_visita(request, visita_id):
    """
    Permite registrar rápidamente aprendices de una ficha en una visita.
    Automáticamente trae sus datos y documentos.
    """
    correo, _ = get_sesion_instructor(request)
    owner_user = _obtener_propietario_instructor(request)
    visita = get_object_or_404(
        VisitaInterna, pk=visita_id, correo_responsable__iexact=correo
    )

    if _solicitud_final_historica_interna(visita):
        messages.error(
            request,
            "La solicitud final ya fue enviada previamente. No se permite registrar nuevos aprendices.",
        )
        return redirect("panel_instructor_interno:detalle_visita", pk=visita_id)

    try:
        ficha = Ficha.objects.get(numero=visita.numero_ficha, creado_por=owner_user)
    except Ficha.DoesNotExist:
        messages.error(request, "❌ No se encontró la ficha asociada a esta visita.")
        return redirect("panel_instructor_interno:detalle_visita", pk=visita_id)

    if request.method == "POST":
        aprendices_ids = request.POST.getlist("aprendices[]")
        aprendices_registrados = 0
        aprendices_duplicados = 0
        aprendices_sin_doc_salud = 0

        for aprendiz_id in aprendices_ids:
            try:
                aprendiz = get_object_or_404(Aprendiz, pk=aprendiz_id, ficha=ficha)

                docs_aprendiz = DocumentoSubidoAprendiz.objects.filter(
                    aprendiz=aprendiz
                ).select_related("documento_requerido")

                tiene_doc_salud = bool(aprendiz.documento_adicional) or docs_aprendiz.filter(
                    documento_requerido__categoria=CATEGORIA_DOC_SALUD
                ).exists()
                if not tiene_doc_salud:
                    aprendices_sin_doc_salud += 1
                    continue

                if AsistenteVisitaInterna.objects.filter(
                    visita=visita, numero_documento=aprendiz.numero_documento
                ).exists():
                    aprendices_duplicados += 1
                    continue

                asistente = AsistenteVisitaInterna.objects.create(
                    visita=visita,
                    nombre_completo=aprendiz.get_nombre_completo(),
                    tipo_documento=aprendiz.tipo_documento,
                    numero_documento=aprendiz.numero_documento,
                    correo=aprendiz.correo,
                    telefono=aprendiz.telefono,
                    estado="pendiente_documentos",
                )

                for doc_subido in docs_aprendiz:
                    DocumentoSubidoAsistente.objects.update_or_create(
                        documento_requerido=doc_subido.documento_requerido,
                        asistente_interna=asistente,
                        defaults={"archivo": doc_subido.archivo},
                    )

                if not docs_aprendiz.exists() and aprendiz.documento_adicional:
                    doc_salud = Documento.objects.filter(
                        categoria="Formato Auto Reporte Condiciones de Salud"
                    ).first()
                    if doc_salud:
                        DocumentoSubidoAsistente.objects.update_or_create(
                            documento_requerido=doc_salud,
                            asistente_interna=asistente,
                            defaults={"archivo": aprendiz.documento_adicional},
                        )

                aprendices_registrados += 1
            except Exception as e:
                messages.warning(request, f"⚠️ Error al registrar aprendiz: {str(e)}")

        if aprendices_registrados > 0:
            messages.success(
                request,
                f"✅ {aprendices_registrados} aprendiz(ces) registrado(s) en la visita con documentos.",
            )
        if aprendices_duplicados > 0:
            messages.warning(
                request,
                f"⚠️ {aprendices_duplicados} aprendiz(ces) ya estaban registrados.",
            )
        if aprendices_sin_doc_salud > 0:
            messages.warning(
                request,
                f"⚠️ {aprendices_sin_doc_salud} aprendiz(ces) no se registraron porque no tienen cargado el reporte de salud.",
            )

        return redirect("panel_instructor_interno:detalle_visita", pk=visita_id)

    aprendices = ficha.aprendices.filter(estado="activo").order_by("apellido", "nombre")

    for aprendiz in aprendices:
        docs_qs = aprendiz.documentos_subidos.select_related("documento_requerido")
        aprendiz.tiene_documentos_subidos = docs_qs.exists()
        aprendiz.tiene_doc_salud = bool(aprendiz.documento_adicional) or docs_qs.filter(
            documento_requerido__categoria="Formato Auto Reporte Condiciones de Salud"
        ).exists()

    ya_registrados = list(
        AsistenteVisitaInterna.objects.filter(visita=visita).values_list(
            "numero_documento", flat=True
        )
    )

    context = {
        "visita": visita,
        "ficha": ficha,
        "aprendices": aprendices,
        "ya_registrados": ya_registrados,
        "total_aprendices": aprendices.count(),
        "aprendices_registrados": AsistenteVisitaInterna.objects.filter(
            visita=visita
        ).count(),
        "correo": correo,
    }

    return render(
        request, "panel_instructor_interno/registrar_aprendices_visita.html", context
    )




def obtener_documentos_por_categoria():
    docs = Documento.objects.all()
    resultado = {}
    for d in docs:
        if d.archivo:
            cat_display = d.get_categoria_display() if d.categoria else "General"
            if cat_display not in resultado:
                resultado[cat_display] = []
            resultado[cat_display].append(d)
    return resultado


def guardar_documentos_aprendiz(aprendiz, archivos_subidos, documentos_por_categoria):
    """Guarda o reemplaza los documentos requeridos cargados para un aprendiz."""
    for docs in documentos_por_categoria.values():
        for doc in docs:
            archivo = archivos_subidos.get(f"documento_{doc.id}")
            if not archivo:
                continue

            DocumentoSubidoAprendiz.objects.update_or_create(
                aprendiz=aprendiz,
                documento_requerido=doc,
                defaults={
                    "archivo": archivo,
                    "estado": "pendiente",
                    "observaciones_revision": None,
                },
            )


def _obtener_docs_subidos_ids_aprendiz(aprendiz, documentos_por_categoria):
    """Retorna IDs de documentos ya cargados, incluyendo compatibilidad con campos legacy."""
    docs_ids = set(
        aprendiz.documentos_subidos.values_list("documento_requerido_id", flat=True)
    )

    if getattr(aprendiz, "documento_adicional", None):
        for categoria, docs in documentos_por_categoria.items():
            cat_norm = _normalizar_categoria_texto(categoria)
            if "auto reporte condiciones de salud" not in cat_norm:
                continue
            for doc in docs:
                docs_ids.add(doc.id)

    return docs_ids


def validar_carga_documentos_aprendiz(
    archivos_subidos,
    documentos_por_categoria,
    tipo_documento,
    docs_subidos_ids=None,
):
    """Valida extension, peso y obligatoriedad de documentos dinamicos por tipo de documento."""
    errores = []
    docs_subidos_ids = set(docs_subidos_ids or [])
    _tipo_doc = str(tipo_documento or "").strip().upper()

    for campo, archivo in archivos_subidos.items():
        if not campo.startswith("documento_") or not archivo:
            continue

        extension = os.path.splitext(archivo.name)[1].lower()
        if extension not in EXTENSIONES_DOCUMENTOS_APRENDIZ:
            errores.append(
                f'El archivo "{archivo.name}" no es valido. Solo se permiten PDF o Word (.doc, .docx).'
            )

        if getattr(archivo, "size", 0) > MAX_TAMANO_DOCUMENTO_APRENDIZ:
            errores.append(
                f'El archivo "{archivo.name}" supera el tamano maximo permitido de 10MB.'
            )

    for categoria, docs in documentos_por_categoria.items():
        cat_norm = _normalizar_categoria_texto(categoria)
        es_doc_salud = "auto reporte condiciones de salud" in cat_norm

        if not es_doc_salud:
            continue

        for doc in docs:
            nombre_campo = f"documento_{doc.id}"
            ya_cargado = doc.id in docs_subidos_ids
            subido_en_post = bool(archivos_subidos.get(nombre_campo))

            if ya_cargado or subido_en_post:
                continue

            if es_doc_salud:
                errores.append(
                    f'Falta cargar el documento obligatorio "{doc.titulo}" (Auto Reporte de Salud).'
                )

    return errores




@instructor_interno_required
def crear_aprendiz(request, ficha_id):
    correo, _ = get_sesion_instructor(request)
    owner_user = _obtener_propietario_instructor(request)
    ficha = get_object_or_404(Ficha, pk=ficha_id, creado_por=owner_user)
    es_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"

    docs_cat = obtener_documentos_por_categoria()

    if request.method == "POST":
        form = AprendizForm(request.POST, request.FILES, ficha=ficha)
        errores_docs_dinamicos = validar_carga_documentos_aprendiz(
            archivos_subidos=request.FILES,
            documentos_por_categoria=docs_cat,
            tipo_documento=request.POST.get("tipo_documento"),
            docs_subidos_ids=[],
        )

        if form.is_valid() and not errores_docs_dinamicos:
            aprendiz = form.save(commit=False)
            aprendiz.ficha = ficha
            try:
                aprendiz.save()
                guardar_documentos_aprendiz(aprendiz, request.FILES, docs_cat)
                mensaje_ok = (
                    f'✅ Aprendiz "{aprendiz.get_nombre_completo()}" registrado correctamente.'
                )

                if es_ajax:
                    return JsonResponse(
                        {
                            "success": True,
                            "message": mensaje_ok,
                            "redirect_url": reverse(
                                "panel_instructor_interno:detalle_aprendices_ficha",
                                kwargs={"pk": ficha.id},
                            ),
                            "aprendiz": {
                                "id": aprendiz.id,
                                "nombre_completo": aprendiz.get_nombre_completo(),
                                "tipo_documento": aprendiz.tipo_documento,
                                "tipo_documento_display": aprendiz.get_tipo_documento_display(),
                                "numero_documento": aprendiz.numero_documento,
                                "correo": aprendiz.correo,
                                "telefono": aprendiz.telefono or "",
                                "estado": aprendiz.estado,
                                "estado_display": aprendiz.get_estado_display(),
                                "edit_url": reverse(
                                    "panel_instructor_interno:editar_aprendiz",
                                    kwargs={"pk": aprendiz.id},
                                ),
                                "delete_url": reverse(
                                    "panel_instructor_interno:eliminar_aprendiz",
                                    kwargs={"pk": aprendiz.id},
                                ),
                            },
                        }
                    )

                messages.success(
                    request,
                    mensaje_ok,
                )
                return redirect(
                    "panel_instructor_interno:detalle_aprendices_ficha", pk=ficha.id
                )
            except IntegrityError:
                mensaje_error = (
                    f"❌ Ya existe un aprendiz con el documento {aprendiz.numero_documento} en esta ficha."
                )
                if es_ajax:
                    return JsonResponse(
                        {
                            "success": False,
                            "error": mensaje_error,
                        },
                        status=400,
                    )
                messages.error(
                    request,
                    mensaje_error,
                )
        else:
            errores_validacion = []
            for error in errores_docs_dinamicos:
                errores_validacion.append(str(error))
            for field, errors in form.errors.items():
                for error in errors:
                    errores_validacion.append(f"{field}: {error}")

            if es_ajax:
                return JsonResponse(
                    {
                        "success": False,
                        "error": "No fue posible registrar el aprendiz. Revisa los campos e intenta nuevamente.",
                        "errores": errores_validacion,
                    },
                    status=400,
                )

            for error in errores_validacion:
                messages.error(request, f"❌ {error}")
    else:
        form = AprendizForm(ficha=ficha)

    context = {
        "form": form,
        "ficha": ficha,
        "correo": correo,
        "titulo": "Registrar Aprendiz",
        "documentos_por_categoria": docs_cat,  # <-- ESTO ACTIVA LAS SECCIONES EN EL HTML
        "docs_subidos_ids": [],
        "es_embed": request.GET.get("embed") == "1",
    }
    return render(request, "panel_instructor_interno/form_aprendiz.html", context)


@instructor_interno_required
def editar_aprendiz(request, pk):
    correo, _ = get_sesion_instructor(request)
    owner_user = _obtener_propietario_instructor(request)
    aprendiz = get_object_or_404(Aprendiz, pk=pk, ficha__creado_por=owner_user)
    ficha = aprendiz.ficha

    docs_cat = obtener_documentos_por_categoria()

    if request.method == "POST":
        docs_subidos_ids = _obtener_docs_subidos_ids_aprendiz(
            aprendiz, docs_cat
        )
        form = AprendizForm(request.POST, request.FILES, instance=aprendiz, ficha=ficha)
        errores_docs_dinamicos = validar_carga_documentos_aprendiz(
            archivos_subidos=request.FILES,
            documentos_por_categoria=docs_cat,
            tipo_documento=request.POST.get("tipo_documento"),
            docs_subidos_ids=docs_subidos_ids,
        )

        if form.is_valid() and not errores_docs_dinamicos:
            try:
                form.save()
                guardar_documentos_aprendiz(aprendiz, request.FILES, docs_cat)
                messages.success(
                    request,
                    f'✅ Aprendiz "{aprendiz.get_nombre_completo()}" actualizado.',
                )
                return redirect(
                    "panel_instructor_interno:detalle_aprendices_ficha", pk=ficha.id
                )
            except IntegrityError:
                messages.error(
                    request,
                    f"❌ Ya existe un aprendiz con el documento {aprendiz.numero_documento} en esta ficha.",
                )
        else:
            for error in errores_docs_dinamicos:
                messages.error(request, f"❌ {error}")
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"❌ {field}: {error}")
    else:
        form = AprendizForm(instance=aprendiz, ficha=ficha)

    docs_subidos_ids = list(_obtener_docs_subidos_ids_aprendiz(aprendiz, docs_cat))

    context = {
        "form": form,
        "aprendiz": aprendiz,
        "ficha": ficha,
        "correo": correo,
        "titulo": "Editar Aprendiz",
        "documentos_por_categoria": docs_cat,  # <-- ESTO ACTIVA LAS SECCIONES EN EL HTML
        "docs_subidos_ids": docs_subidos_ids,
    }
    return render(request, "panel_instructor_interno/form_aprendiz.html", context)


@instructor_interno_required
def eliminar_asistente_visita(request, visita_id, asistente_id, tipo):
    """
    Elimina un asistente desde el detalle de una visita del instructor.
    Redirige de vuelta a la visita, no a visitante.
    """
    correo, _ = get_sesion_instructor(request)
    visita = get_object_or_404(
        VisitaInterna, pk=visita_id, correo_responsable__iexact=correo
    )

    if _solicitud_final_historica_interna(visita):
        messages.error(
            request,
            "La solicitud final ya fue enviada previamente. No es posible eliminar aprendices.",
        )
        return redirect("panel_instructor_interno:detalle_visita", pk=visita_id)

    if tipo == "interna":
        asistente = get_object_or_404(
            AsistenteVisitaInterna, pk=asistente_id, visita=visita
        )
        nombre = asistente.nombre_completo
        asistente.delete()
    else:
        return redirect("panel_instructor_interno:detalle_visita", pk=visita_id)

    messages.success(request, f'🗑️ Asistente "{nombre}" eliminado de la visita.')
    return redirect("panel_instructor_interno:detalle_visita", pk=visita_id)
