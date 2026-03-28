from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.contrib.auth.tokens import default_token_generator
from django.utils.encoding import force_bytes, force_str
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.views.decorators.csrf import csrf_protect
from django.urls import reverse
from datetime import datetime, timedelta
import random
from visitaInterna.models import (
    VisitaInterna,
    AsistenteVisitaInterna,
    HistorialAccionVisitaInterna,
)
from visitaExterna.models import (
    VisitaExterna,
    AsistenteVisitaExterna,
    HistorialAccionVisitaExterna,
)
from documentos.models import (
    Documento,
    DocumentoSubidoAsistente,
    DocumentoSubidoAprendiz,
)
from .forms import (
    RegistroVisitanteForm,
    PasswordResetRequestForm,
    PasswordResetConfirmForm,
)
from .forms import (
    RegistroVisitanteForm,
    PasswordResetRequestForm,
    PasswordResetConfirmForm,
    ActualizarPerfilForm,
    CambiarContrasenaForm,
    LoginResponsableForm,
    VerificacionCodigoRegistroForm,
)
from .models import RegistroVisitante
from django.conf import settings
from django.db import IntegrityError, transaction
from pathlib import Path
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
import re
import os
import csv
import io


CATEGORIAS_ARCHIVOS_FINALES = {
    "ats",
    "formato induccion y reinduccion",
    "charla de seguridad y calestenia",
    "charla de seguridad y calistenia",
}
CATEGORIAS_DOCUMENTOS_REGISTRO = {
    "formato auto reporte condiciones de salud",
    "formato autorizacion padres de familia",
}
CATEGORIA_AUTORIZACION_PADRES = "Formato Autorización Padres de Familia"
AUTH_VISITANTE_MESSAGE_TAG = "auth_visitante"
REGISTRO_VERIFICACION_SESSION_KEY = "registro_verificacion_pendiente"
REGISTRO_VERIFICACION_TTL_MINUTOS = 10
FORMATOS_CARGA_MASIVA_ASISTENTES = {".csv", ".xlsx", ".xls", ".pdf"}
MARCADOR_SOLICITUD_FINAL_ENVIADA = "[SOLICITUD_FINAL_ENVIADA]"
COLUMNAS_CARGA_MASIVA_ASISTENTES = [
    "nombre_completo",
    "tipo_documento",
    "numero_documento",
    "correo",
    "telefono",
]
EJEMPLOS_CARGA_MASIVA_ASISTENTES = [
    {
        "nombre_completo": "Laura Gomez",
        "tipo_documento": "CC",
        "numero_documento": "1022334455",
        "correo": "laura.gomez@example.com",
        "telefono": "3001234567",
    },
    {
        "nombre_completo": "Mateo Rodriguez",
        "tipo_documento": "TI",
        "numero_documento": "1099988877",
        "correo": "mateo.rodriguez@example.com",
        "telefono": "3017654321",
    },
]
MAX_INTENTOS_LOGIN = 5
MINUTOS_BLOQUEO_LOGIN = 10


def _generar_codigo_verificacion(longitud=6):
    generador = random.SystemRandom()
    return "".join(str(generador.randint(0, 9)) for _ in range(longitud))


def _normalizar_encabezado_importacion(valor):
    return (
        str(valor or "")
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )


def _mapear_tipo_documento_importacion(valor):
    tipo = str(valor or "").strip().upper()
    equivalencias = {
        "CEDULA": "CC",
        "CEDULA_DE_CIUDADANIA": "CC",
        "TARJETA_DE_IDENTIDAD": "TI",
        "PASAPORTE": "PP",
    }
    return equivalencias.get(tipo, tipo)


def _extraer_filas_csv_asistentes(archivo):
    contenido = archivo.read()
    if isinstance(contenido, bytes):
        try:
            texto = contenido.decode("utf-8-sig")
        except UnicodeDecodeError:
            texto = contenido.decode("latin-1")
    else:
        texto = str(contenido)

    lector = csv.DictReader(io.StringIO(texto))
    if not lector.fieldnames:
        return [], ["El CSV no contiene encabezados."]

    headers = [_normalizar_encabezado_importacion(h) for h in lector.fieldnames]
    faltantes = [c for c in COLUMNAS_CARGA_MASIVA_ASISTENTES if c not in headers]
    if faltantes:
        return [], [
            "La plantilla no coincide. Faltan columnas: " + ", ".join(faltantes)
        ]

    filas = []
    for idx, row in enumerate(lector, start=2):
        fila = {}
        for key, value in row.items():
            fila[_normalizar_encabezado_importacion(key)] = (value or "").strip()
        if not any(fila.values()):
            continue
        filas.append((idx, fila))

    if not filas:
        return [], ["No se encontraron filas de asistentes para importar."]

    return filas, []


def _extraer_filas_excel_asistentes(archivo):
    try:
        from openpyxl import load_workbook
    except Exception:
        return [], [
            "No se pudo procesar Excel. Instale openpyxl para habilitar este formato."
        ]

    try:
        wb = load_workbook(filename=archivo, data_only=True)
        ws = wb.active
    except Exception:
        return [], ["No fue posible leer el archivo Excel. Verifique la plantilla."]

    encabezados = []
    for cell in ws[1]:
        encabezados.append(_normalizar_encabezado_importacion(cell.value))

    faltantes = [c for c in COLUMNAS_CARGA_MASIVA_ASISTENTES if c not in encabezados]
    if faltantes:
        return [], [
            "La plantilla no coincide. Faltan columnas: " + ", ".join(faltantes)
        ]

    filas = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        fila = {}
        for col_idx, valor in enumerate(row):
            if col_idx >= len(encabezados):
                continue
            clave = encabezados[col_idx]
            fila[clave] = "" if valor is None else str(valor).strip()

        if not any(fila.values()):
            continue
        filas.append((idx, fila))

    if not filas:
        return [], ["No se encontraron filas de asistentes para importar."]

    return filas, []


def _extraer_filas_pdf_asistentes(archivo):
    try:
        from pypdf import PdfReader
    except Exception:
        return [], ["No se pudo procesar PDF. Instale pypdf para habilitar este formato."]

    try:
        reader = PdfReader(archivo)
        texto = "\n".join((p.extract_text() or "") for p in reader.pages)
    except Exception:
        return [], ["No fue posible leer el archivo PDF. Use la plantilla descargable."]

    filas = []
    for idx, linea in enumerate(texto.splitlines(), start=1):
        normalizada = linea.strip()
        if not normalizada or ";" not in normalizada:
            continue
        partes = [p.strip() for p in normalizada.split(";")]
        if len(partes) < 5:
            continue
        if _normalizar_encabezado_importacion(partes[0]) == "nombre_completo":
            continue

        filas.append(
            (
                idx,
                {
                    "nombre_completo": partes[0],
                    "tipo_documento": partes[1],
                    "numero_documento": partes[2],
                    "correo": partes[3],
                    "telefono": partes[4],
                },
            )
        )

    if not filas:
        return [], [
            "No se detectaron filas válidas en el PDF. Mantenga el formato con ';'."
        ]

    return filas, []


def _extraer_filas_importacion_asistentes(archivo):
    extension = os.path.splitext(getattr(archivo, "name", ""))[1].lower()

    if extension not in FORMATOS_CARGA_MASIVA_ASISTENTES:
        return [], ["Formato no permitido. Use CSV, Excel (.xlsx/.xls) o PDF."]

    if extension == ".csv":
        return _extraer_filas_csv_asistentes(archivo)

    if extension in {".xlsx", ".xls"}:
        return _extraer_filas_excel_asistentes(archivo)

    return _extraer_filas_pdf_asistentes(archivo)


def _validar_fila_asistente_importacion(fila):
    errores = []
    tipos_documento_validos = {"CC", "CE", "TI", "PPT", "PP"}

    nombre = " ".join(str(fila.get("nombre_completo", "")).split())
    tipo_doc = _mapear_tipo_documento_importacion(fila.get("tipo_documento"))
    num_doc = str(fila.get("numero_documento", "")).strip()
    correo = str(fila.get("correo", "")).strip().lower()
    telefono = str(fila.get("telefono", "")).strip()

    if not nombre or len(nombre) < 3 or len(nombre) > 80:
        errores.append("nombre_completo invalido (3 a 80 caracteres).")
    elif not re.fullmatch(r"[A-Za-zÁÉÍÓÚáéíóúÑñÜü\s]+", nombre):
        errores.append("nombre_completo solo puede contener letras y espacios.")

    if tipo_doc not in tipos_documento_validos:
        errores.append("tipo_documento no valido.")

    if not num_doc or not num_doc.isdigit() or not 5 <= len(num_doc) <= 10:
        errores.append("numero_documento debe tener entre 5 y 10 digitos.")

    if correo:
        try:
            validate_email(correo)
        except ValidationError:
            errores.append("correo invalido.")

    if telefono:
        if not telefono.isdigit() or not 7 <= len(telefono) <= 10:
            errores.append("telefono debe tener entre 7 y 10 digitos.")

    return {
        "nombre_completo": nombre,
        "tipo_documento": tipo_doc,
        "numero_documento": num_doc,
        "correo": correo,
        "telefono": telefono,
    }, errores


def _sincronizar_asistente_con_ficha_interna(visita, datos_asistente):
    try:
        from panel_instructor_interno.models import Ficha, Aprendiz
    except Exception:
        return

    ficha = Ficha.objects.filter(numero=visita.numero_ficha).first()
    if not ficha:
        return

    existe = Aprendiz.objects.filter(
        ficha=ficha,
        numero_documento=datos_asistente["numero_documento"],
    ).exists()
    if existe:
        return

    partes_nombre = datos_asistente["nombre_completo"].split()
    nombre = partes_nombre[0] if partes_nombre else ""
    apellido = " ".join(partes_nombre[1:]) if len(partes_nombre) > 1 else ""

    Aprendiz.objects.create(
        ficha=ficha,
        nombre=nombre,
        apellido=apellido,
        tipo_documento=datos_asistente["tipo_documento"],
        numero_documento=datos_asistente["numero_documento"],
        correo=datos_asistente["correo"],
        telefono=datos_asistente["telefono"],
        estado="activo",
    )


def _procesar_carga_masiva_asistentes(visita, tipo, filas, max_asistentes):
    creados = 0
    duplicados = 0
    omitidos_limite = 0
    errores = []

    modelo_asistente = AsistenteVisitaInterna if tipo == "interna" else AsistenteVisitaExterna
    total_actual = visita.asistentes.count()

    for fila_numero, fila in filas:
        if total_actual + creados >= max_asistentes:
            omitidos_limite += 1
            continue

        datos, errores_validacion = _validar_fila_asistente_importacion(fila)
        if errores_validacion:
            errores.append(f"Fila {fila_numero}: {' | '.join(errores_validacion)}")
            continue

        if modelo_asistente.objects.filter(
            visita=visita,
            numero_documento=datos["numero_documento"],
        ).exists():
            duplicados += 1
            continue

        asistente = modelo_asistente.objects.create(
            visita=visita,
            nombre_completo=datos["nombre_completo"],
            tipo_documento=datos["tipo_documento"],
            numero_documento=datos["numero_documento"],
            correo=datos["correo"],
            telefono=datos["telefono"],
        )

        if tipo == "interna":
            _sincronizar_asistente_con_ficha_interna(visita, datos)

        asistente.tiene_doc_salud = False
        creados += 1

    return {
        "creados": creados,
        "duplicados": duplicados,
        "omitidos_limite": omitidos_limite,
        "errores": errores,
    }


def _enviar_codigo_verificacion_registro(correo, codigo, nombre):
    asunto = "Codigo de verificacion de cuenta - MINEGATE"
    html_content = render_to_string(
        "emails/codigo_verificacion_registro.html",
        {
            "nombre": nombre,
            "codigo": codigo,
            "minutos": REGISTRO_VERIFICACION_TTL_MINUTOS,
        },
    )
    text_content = strip_tags(html_content)

    email = EmailMultiAlternatives(
        subject=asunto,
        body=text_content,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[correo],
    )
    email.attach_alternative(html_content, "text/html")
    email.send()


def _redirect_segun_rol(request, tipo=None, visita_id=None):
    """
    Redirige al panel correcto según el rol de la sesión.
    Si se pasa tipo y visita_id, intenta ir al detalle de la visita.
    """
    rol = request.session.get("responsable_rol")
    if rol == "interno":
        if tipo and visita_id:
            from django.urls import reverse

            return redirect(
                reverse("panel_instructor_interno:detalle_visita", args=[visita_id])
            )
        return redirect("panel_instructor_interno:panel")
    elif rol == "externo":
        if tipo and visita_id:
            from django.urls import reverse

            return redirect(
                reverse("panel_instructor_externo:detalle_visita", args=[visita_id])
            )
        return redirect("panel_instructor_externo:panel")
    else:
        return redirect("panel_visitante:panel_responsable")


def _tiene_acceso_por_correo(visita, correo):
    """
    Valida ownership de la visita usando el correo del responsable.
    El correo es el identificador canonico de la cuenta en sesion.
    """
    if not correo:
        return False
    return (visita.correo_responsable or "").strip().lower() == correo.strip().lower()


def _solicitud_final_ya_enviada(visita, tipo):
    """
    Indica si la visita ya fue enviada a revision al menos una vez.
    Se considera historico, incluso si el estado volvio a aprobada_inicial
    por una devolucion de correcciones.
    """
    if visita.estado in ["documentos_enviados", "en_revision_documentos", "confirmada"]:
        return True

    if tipo == "interna":
        if HistorialAccionVisitaInterna.objects.filter(
            visita=visita,
            descripcion__icontains=MARCADOR_SOLICITUD_FINAL_ENVIADA,
        ).exists():
            return True
        if HistorialAccionVisitaInterna.objects.filter(
            visita=visita,
            tipo_accion__in=["inicio_revision", "devolucion_correccion", "confirmacion"],
        ).exists():
            return True
    else:
        if HistorialAccionVisitaExterna.objects.filter(
            visita=visita,
            descripcion__icontains=MARCADOR_SOLICITUD_FINAL_ENVIADA,
        ).exists():
            return True
        if HistorialAccionVisitaExterna.objects.filter(
            visita=visita,
            tipo_accion__in=["inicio_revision", "devolucion_correccion", "confirmacion"],
        ).exists():
            return True

    if visita.asistentes.filter(
        estado__in=["documentos_aprobados", "documentos_rechazados"]
    ).exists():
        return True

    filtros_docs = {"estado__in": ["aprobado", "rechazado"]}
    if tipo == "interna":
        filtros_docs["asistente_interna__visita"] = visita
    else:
        filtros_docs["asistente_externa__visita"] = visita

    return DocumentoSubidoAsistente.objects.filter(**filtros_docs).exists()


def _normalizar_categoria_texto(valor):
    return (
        str(valor or "")
        .strip()
        .lower()
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )


def _es_categoria_archivo_final(categoria):
    return _normalizar_categoria_texto(categoria) in CATEGORIAS_ARCHIVOS_FINALES


def _es_categoria_documento_registro(categoria):
    return _normalizar_categoria_texto(categoria) in CATEGORIAS_DOCUMENTOS_REGISTRO


def _agrupar_documentos_por_categoria(documentos):
    documentos_por_categoria = {}
    for doc in documentos:
        categoria = doc.categoria
        if categoria not in documentos_por_categoria:
            documentos_por_categoria[categoria] = []
        documentos_por_categoria[categoria].append(doc)
    return documentos_por_categoria


def _documentos_actuales_asistente(asistente, tipo):
    filtros = (
        {"asistente_interna": asistente}
        if tipo == "interna"
        else {"asistente_externa": asistente}
    )

    docs = (
        DocumentoSubidoAsistente.objects.filter(**filtros)
        .select_related("documento_requerido")
        .order_by("documento_requerido_id", "-fecha_subida", "-id")
    )

    latest_por_documento = {}
    for ds in docs:
        if ds.documento_requerido_id not in latest_por_documento:
            latest_por_documento[ds.documento_requerido_id] = ds

    return list(latest_por_documento.values())


def _sincronizar_estado_asistente_por_docs(asistente, tipo):
    docs_actuales = _documentos_actuales_asistente(asistente, tipo)
    docs_personales = [
        ds
        for ds in docs_actuales
        if not _es_categoria_archivo_final(ds.documento_requerido.categoria)
    ]

    tiene_rechazos = any(ds.estado == "rechazado" for ds in docs_personales)
    todos_aprobados = bool(docs_personales) and all(
        ds.estado == "aprobado" for ds in docs_personales
    )

    if getattr(asistente, "formato_autorizacion_padres", None):
        estado_autorizacion = getattr(
            asistente, "estado_autorizacion_padres", "pendiente"
        )
        if estado_autorizacion == "rechazado":
            tiene_rechazos = True
        elif estado_autorizacion != "aprobado":
            todos_aprobados = False

    if tiene_rechazos:
        nuevo_estado = "documentos_rechazados"
    elif todos_aprobados:
        nuevo_estado = "documentos_aprobados"
    else:
        nuevo_estado = "pendiente_documentos"

    update_fields = []
    if asistente.estado != nuevo_estado:
        asistente.estado = nuevo_estado
        update_fields.append("estado")

    if nuevo_estado != "documentos_rechazados" and asistente.observaciones_revision:
        asistente.observaciones_revision = ""
        update_fields.append("observaciones_revision")

    if update_fields:
        asistente.save(update_fields=update_fields)

    return nuevo_estado


def _resumen_pendientes_correccion(visita, tipo):
    asistentes_rechazados = visita.asistentes.filter(
        estado="documentos_rechazados"
    ).count()

    docs_finales_requeridos = Documento.objects.filter(
        categoria__in=[
            "ATS",
            "Formato Inducción y Reinducción",
            "Charla de Seguridad y Calestenia",
        ]
    )

    archivos_finales_rechazados = 0
    archivos_finales_faltantes = 0

    for doc_final in docs_finales_requeridos:
        filtros = {"documento_requerido": doc_final}
        if tipo == "interna":
            filtros["asistente_interna__visita"] = visita
        else:
            filtros["asistente_externa__visita"] = visita

        ultimo_archivo = (
            DocumentoSubidoAsistente.objects.filter(**filtros)
            .order_by("-fecha_subida", "-id")
            .first()
        )

        if not ultimo_archivo:
            archivos_finales_faltantes += 1
        elif ultimo_archivo.estado == "rechazado":
            archivos_finales_rechazados += 1

    pendientes_correccion = (
        asistentes_rechazados + archivos_finales_rechazados + archivos_finales_faltantes
    )

    return {
        "asistentes_rechazados": asistentes_rechazados,
        "archivos_finales_rechazados": archivos_finales_rechazados,
        "archivos_finales_faltantes": archivos_finales_faltantes,
        "pendientes_correccion": pendientes_correccion,
    }


def _auto_reenviar_a_revision_si_aplica(visita, tipo, resumen_pendientes):
    """Reenvia automaticamente a revision cuando una correccion historica queda completa."""
    if resumen_pendientes.get("pendientes_correccion", 0) != 0:
        return False

    if visita.estado != "aprobada_inicial":
        return False

    if not _solicitud_final_ya_enviada(visita, tipo):
        return False

    visita.estado = "documentos_enviados"
    visita.save(update_fields=["estado"])
    return True


def login_responsable(request):
    """
    Login para responsables de visitas usando solo documento y contraseña.
    """
    form = LoginResponsableForm(request.POST or None)

    if request.method == "POST":
        if not form.is_valid():
            return render(request, "login_responsable.html", {"form": form})

        documento = form.cleaned_data["documento"]
        contrasena = form.cleaned_data["contrasena"]

        visitante = RegistroVisitante.objects.filter(documento=documento).first()

        if visitante and visitante.bloqueado_hasta:
            ahora = timezone.now()
            if visitante.bloqueado_hasta <= ahora:
                visitante.bloqueado_hasta = None
                visitante.intentos_fallidos = 0
                visitante.save(update_fields=["bloqueado_hasta", "intentos_fallidos"])
            else:
                segundos_restantes = int((visitante.bloqueado_hasta - ahora).total_seconds())
                minutos_restantes = max(1, (segundos_restantes + 59) // 60)
                messages.error(
                    request,
                    f"Usuario bloqueado temporalmente. Intenta de nuevo en {minutos_restantes} minuto(s).",
                    extra_tags=AUTH_VISITANTE_MESSAGE_TAG,
                )
                return render(request, "login_responsable.html", {"form": form})

        if visitante and visitante.check_password(contrasena):
            if visitante.intentos_fallidos or visitante.bloqueado_hasta:
                visitante.intentos_fallidos = 0
                visitante.bloqueado_hasta = None
                visitante.save(update_fields=["intentos_fallidos", "bloqueado_hasta"])

            # Guardar sesión
            request.session["responsable_correo"] = visitante.correo
            request.session["responsable_documento"] = visitante.documento
            request.session["responsable_rol"] = visitante.rol
            request.session["responsable_nombre"] = visitante.nombre
            request.session["responsable_apellido"] = visitante.apellido
            request.session["responsable_tipo_documento"] = visitante.tipo_documento
            request.session["responsable_telefono"] = visitante.telefono
            request.session["responsable_autenticado"] = True
            request.session.modified = True

            messages.success(
                request,
                f"Bienvenido {visitante.nombre} {visitante.apellido}. Accediendo a tu panel...",
                extra_tags=AUTH_VISITANTE_MESSAGE_TAG,
            )
            # Redirigir al panel de instructor según el rol
            if visitante.rol == "interno":
                return redirect("panel_instructor_interno:panel")
            else:
                return redirect("panel_instructor_externo:panel")

        if visitante:
            visitante.intentos_fallidos += 1

            if visitante.intentos_fallidos >= MAX_INTENTOS_LOGIN:
                visitante.bloqueado_hasta = timezone.now() + timedelta(
                    minutes=MINUTOS_BLOQUEO_LOGIN
                )
                visitante.save(update_fields=["intentos_fallidos", "bloqueado_hasta"])
                messages.error(
                    request,
                    "Se alcanzaron 5 intentos fallidos. Tu usuario ha sido bloqueado por 10 minutos.",
                    extra_tags=AUTH_VISITANTE_MESSAGE_TAG,
                )
                return render(request, "login_responsable.html", {"form": form})

            visitante.save(update_fields=["intentos_fallidos"])
            intentos_restantes = MAX_INTENTOS_LOGIN - visitante.intentos_fallidos
            messages.error(
                request,
                f"Credenciales invalidas. Verifica tu documento y contrasena. Te quedan {intentos_restantes} intento(s).",
                extra_tags=AUTH_VISITANTE_MESSAGE_TAG,
            )
            return render(request, "login_responsable.html", {"form": form})

        messages.error(
            request,
            "Credenciales invalidas. Verifica tu documento y contrasena.",
            extra_tags=AUTH_VISITANTE_MESSAGE_TAG,
        )

    return render(request, "login_responsable.html", {"form": form})


def registro_visita(request):
    """
    Registro inicial de visita para responsables.
    """
    if request.method == "POST":
        form = RegistroVisitanteForm(request.POST)

        if form.is_valid():
            correo = form.cleaned_data["correo"].strip().lower()
            codigo = _generar_codigo_verificacion()

            request.session[REGISTRO_VERIFICACION_SESSION_KEY] = {
                "nombre": form.cleaned_data["nombre"],
                "apellido": form.cleaned_data["apellido"],
                "tipo_documento": form.cleaned_data["tipo_documento"],
                "documento": form.cleaned_data["documento"],
                "telefono": form.cleaned_data["telefono"],
                "correo": correo,
                "rol": form.cleaned_data["rol"],
                "password1": form.cleaned_data["password1"],
                "codigo": codigo,
                "creado_en": timezone.now().isoformat(),
            }
            request.session.modified = True

            try:
                _enviar_codigo_verificacion_registro(
                    correo=correo,
                    codigo=codigo,
                    nombre=form.cleaned_data["nombre"],
                )
            except Exception:
                messages.error(
                    request,
                    "No fue posible enviar el codigo de verificacion al correo. Intente nuevamente.",
                    extra_tags=AUTH_VISITANTE_MESSAGE_TAG,
                )
                return render(
                    request,
                    "registro_visita.html",
                    {"form": form, "titulo": "Registro de Usuario"},
                )

            messages.info(
                request,
                f"Enviamos un codigo de verificacion a {correo}. Ingreselo para finalizar el registro.",
                extra_tags=AUTH_VISITANTE_MESSAGE_TAG,
            )
            return redirect("panel_visitante:verificar_codigo_registro")
    else:
        form = RegistroVisitanteForm()

    return render(
        request,
        "registro_visita.html",
        {"form": form, "titulo": "Registro de Usuario"},
    )


def verificar_codigo_registro(request):
    datos_registro = request.session.get(REGISTRO_VERIFICACION_SESSION_KEY)
    if not datos_registro:
        messages.warning(
            request,
            "No hay un registro pendiente por verificar.",
            extra_tags=AUTH_VISITANTE_MESSAGE_TAG,
        )
        return redirect("panel_visitante:registro_visita")

    try:
        creado_en = datetime.fromisoformat(datos_registro.get("creado_en", ""))
    except (TypeError, ValueError):
        creado_en = None

    expirado = True
    if creado_en is not None:
        expirado = (timezone.now() - creado_en).total_seconds() > (
            REGISTRO_VERIFICACION_TTL_MINUTOS * 60
        )

    if request.method == "POST" and request.POST.get("accion") == "reenviar":
        codigo = _generar_codigo_verificacion()
        datos_registro["codigo"] = codigo
        datos_registro["creado_en"] = timezone.now().isoformat()
        request.session[REGISTRO_VERIFICACION_SESSION_KEY] = datos_registro
        request.session.modified = True

        try:
            _enviar_codigo_verificacion_registro(
                correo=datos_registro["correo"],
                codigo=codigo,
                nombre=datos_registro["nombre"],
            )
            messages.success(
                request,
                "Se envio un nuevo codigo de verificacion a su correo.",
                extra_tags=AUTH_VISITANTE_MESSAGE_TAG,
            )
        except Exception:
            messages.error(
                request,
                "No fue posible reenviar el codigo. Intente de nuevo.",
                extra_tags=AUTH_VISITANTE_MESSAGE_TAG,
            )
        return redirect("panel_visitante:verificar_codigo_registro")

    form = VerificacionCodigoRegistroForm(request.POST or None)

    if request.method == "POST" and request.POST.get("accion") != "reenviar":
        if expirado:
            form.add_error("codigo", "El codigo ha expirado. Solicite uno nuevo.")
        elif form.is_valid():
            codigo_ingresado = form.cleaned_data["codigo"]
            codigo_real = str(datos_registro.get("codigo", "")).strip()

            if codigo_ingresado != codigo_real:
                form.add_error("codigo", "Codigo incorrecto.")
            else:
                try:
                    with transaction.atomic():
                        visitante = RegistroVisitante(
                            nombre=datos_registro["nombre"],
                            apellido=datos_registro["apellido"],
                            tipo_documento=datos_registro["tipo_documento"],
                            documento=datos_registro["documento"],
                            telefono=datos_registro["telefono"],
                            correo=datos_registro["correo"],
                            rol=datos_registro["rol"],
                        )
                        visitante.set_password(datos_registro["password1"])
                        visitante.save()
                except IntegrityError:
                    form.add_error(
                        None,
                        "No fue posible crear la cuenta porque el correo o documento ya existe.",
                    )
                else:
                    request.session.pop(REGISTRO_VERIFICACION_SESSION_KEY, None)
                    messages.success(
                        request,
                        f"Cuenta creada exitosamente para {visitante.nombre} {visitante.apellido}. Ya puedes iniciar sesion con tus credenciales.",
                        extra_tags=AUTH_VISITANTE_MESSAGE_TAG,
                    )
                    return redirect("panel_visitante:login_responsable")

    context = {
        "form": form,
        "correo": datos_registro.get("correo"),
        "expirado": expirado,
        "minutos": REGISTRO_VERIFICACION_TTL_MINUTOS,
    }
    return render(request, "verificar_codigo_registro.html", context)


def logout_responsable(request):
    """
    Cerrar sesión del responsable.
    """
    request.session.pop("responsable_correo", None)
    request.session.pop("responsable_documento", None)
    request.session.pop("responsable_autenticado", None)
    messages.success(
        request,
        "Sesión cerrada correctamente.",
        extra_tags=AUTH_VISITANTE_MESSAGE_TAG,
    )
    return redirect("panel_visitante:login_responsable")


def panel_responsable(request):
    """
    Panel para que el responsable vea sus visitas y registre asistentes.
    """
    # Obtener datos de sesión
    correo = request.session.get("responsable_correo")
    documento = request.session.get("responsable_documento")
    rol = request.session.get("responsable_rol")
    autenticado = request.session.get("responsable_autenticado", False)

    # Verificar si está autenticado
    if not autenticado or not correo or not documento:
        messages.warning(
            request,
            "Debe iniciar sesión para acceder al panel.",
            extra_tags=AUTH_VISITANTE_MESSAGE_TAG,
        )
        return redirect("panel_visitante:login_responsable")

    try:
        # Obtener visitas del responsable
        visitas_internas = VisitaInterna.objects.filter(
            correo_responsable__iexact=correo
        ).prefetch_related("asistentes")

        visitas_externas = VisitaExterna.objects.filter(
            correo_responsable__iexact=correo
        ).prefetch_related("asistentes")

        context = {
            "visitas_internas": visitas_internas,
            "visitas_externas": visitas_externas,
            "correo": correo,
            "rol": rol,
        }

        return render(request, "panel_responsable.html", context)

    except Exception as e:
        messages.error(
            request,
            f"Error al cargar el panel: {str(e)}",
            extra_tags=AUTH_VISITANTE_MESSAGE_TAG,
        )
        return redirect("panel_visitante:login_responsable")


def descargar_plantilla_carga_masiva_asistentes(request, tipo, visita_id, formato):
    if not request.session.get("responsable_autenticado"):
        messages.warning(
            request,
            "Debe iniciar sesión para acceder.",
            extra_tags=AUTH_VISITANTE_MESSAGE_TAG,
        )
        return redirect("panel_visitante:login_responsable")

    correo = request.session.get("responsable_correo")
    if tipo == "interna":
        get_object_or_404(VisitaInterna, id=visita_id, correo_responsable__iexact=correo)
        messages.error(
            request,
            "La plantilla de carga masiva de asistentes aplica solo para visitas externas.",
        )
        return redirect(
            "panel_visitante:registrar_asistentes", tipo=tipo, visita_id=visita_id
        )
    elif tipo == "externa":
        get_object_or_404(VisitaExterna, id=visita_id, correo_responsable__iexact=correo)
    else:
        messages.error(request, "Tipo de visita no válido.")
        return _redirect_segun_rol(request)

    formato = str(formato or "").lower()

    if formato == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = (
            'attachment; filename="plantilla_carga_masiva_asistentes.csv"'
        )
        writer = csv.DictWriter(response, fieldnames=COLUMNAS_CARGA_MASIVA_ASISTENTES)
        writer.writeheader()
        writer.writerows(EJEMPLOS_CARGA_MASIVA_ASISTENTES)
        return response

    if formato == "xlsx":
        try:
            from openpyxl import Workbook
        except Exception:
            messages.error(
                request,
                "No fue posible generar la plantilla Excel. Instale openpyxl.",
            )
            return redirect(
                "panel_visitante:registrar_asistentes", tipo=tipo, visita_id=visita_id
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla"
        ws.append(COLUMNAS_CARGA_MASIVA_ASISTENTES)
        for ejemplo in EJEMPLOS_CARGA_MASIVA_ASISTENTES:
            ws.append([ejemplo.get(col, "") for col in COLUMNAS_CARGA_MASIVA_ASISTENTES])

        salida = io.BytesIO()
        wb.save(salida)
        salida.seek(0)

        response = HttpResponse(
            salida.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            'attachment; filename="plantilla_carga_masiva_asistentes.xlsx"'
        )
        return response

    if formato == "pdf":
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas

        output = io.BytesIO()
        pdf = canvas.Canvas(output, pagesize=letter)

        y = 760
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(40, y, "Plantilla carga masiva de asistentes")
        y -= 22
        pdf.setFont("Helvetica", 9)
        pdf.drawString(
            40,
            y,
            "Cada fila debe conservar este orden y usar ';' como separador:",
        )
        y -= 14
        pdf.drawString(40, y, "; ".join(COLUMNAS_CARGA_MASIVA_ASISTENTES))
        y -= 22
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(40, y, "Ejemplos:")
        y -= 16
        pdf.setFont("Helvetica", 9)

        for ejemplo in EJEMPLOS_CARGA_MASIVA_ASISTENTES:
            linea = ";".join(
                [str(ejemplo.get(col, "")) for col in COLUMNAS_CARGA_MASIVA_ASISTENTES]
            )
            pdf.drawString(40, y, linea)
            y -= 14

        y -= 6
        pdf.drawString(
            40,
            y,
            "Nota: si usas PDF para importar, mantén una fila por línea con ese formato.",
        )

        pdf.showPage()
        pdf.save()
        output.seek(0)

        response = HttpResponse(output.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = (
            'attachment; filename="plantilla_carga_masiva_asistentes.pdf"'
        )
        return response

    messages.error(request, "Formato de plantilla no válido.")
    return redirect("panel_visitante:registrar_asistentes", tipo=tipo, visita_id=visita_id)


def registrar_asistentes(request, tipo, visita_id):
    """
    Formulario para registrar asistentes a una visita aprobada.
    """
    es_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"

    # Verificar autenticación
    if not request.session.get("responsable_autenticado"):
        messages.warning(
            request,
            "Debe iniciar sesión para acceder.",
            extra_tags=AUTH_VISITANTE_MESSAGE_TAG,
        )
        return redirect("panel_visitante:login_responsable")

    correo = request.session.get("responsable_correo")
    # Obtener la visita según el tipo
    if tipo == "interna":
        visita = get_object_or_404(
            VisitaInterna,
            id=visita_id,
            correo_responsable__iexact=correo,
        )
        asistentes = visita.asistentes.all()
        max_asistentes = visita.cantidad_aprendices
    elif tipo == "externa":
        visita = get_object_or_404(
            VisitaExterna,
            id=visita_id,
            correo_responsable__iexact=correo,
        )
        asistentes = visita.asistentes.all()
        max_asistentes = visita.cantidad_visitantes
    else:
        messages.error(request, "Tipo de visita no válido.")
        return _redirect_segun_rol(request)

    # Validación de seguridad: la visita debe tener al menos 1 asistente requerido
    if max_asistentes < 1:
        messages.error(
            request,
            "Error: Esta visita debe tener una cantidad válida de asistentes (mínimo 1). "
            "No puede proceder con el registro.",
        )
        return _redirect_segun_rol(request)

    # Verificar que la visita esté aprobada inicialmente para registro de asistentes
    if visita.estado not in [
        "aprobada_inicial",
        "documentos_enviados",
        "en_revision_documentos",
    ]:
        messages.error(request, "Solo puede registrar asistentes en visitas aprobadas.")
        return _redirect_segun_rol(request)

    # Obtener asistentes previos disponibles (de visitas pasadas del mismo programa)
    asistentes_previos = []
    if tipo == "interna":
        # Obtener todas las fichas anteriores del mismo programa
        visitas_anteriores = VisitaInterna.objects.filter(
            numero_ficha=visita.numero_ficha,
            id__lt=visita.id,
            estado__in=["documentos_enviados", "en_revision_documentos", "confirmada"],
        ).order_by("-id")
        for visita_anterior in visitas_anteriores:
            asistentes_del_programa = visita_anterior.asistentes.filter(
                puede_reutilizar=True
            ).values_list(
                "id",
                "nombre_completo",
                "tipo_documento",
                "numero_documento",
                "correo",
                "telefono",
            )
            asistentes_previos.extend(asistentes_del_programa)
        # Eliminar duplicados basados en número de documento
        nums_doc_vistos = set()
        asistentes_previos_unicos = []
        for doc_info in asistentes_previos:
            num_doc = doc_info[3]  # numero_documento
            if num_doc not in nums_doc_vistos:
                nums_doc_vistos.add(num_doc)
                asistentes_previos_unicos.append(doc_info)
        asistentes_previos = asistentes_previos_unicos
    elif tipo == "externa":
        # Obtener todas las visitas anteriores de la misma institución
        visitas_anteriores = VisitaExterna.objects.filter(
            nombre=visita.nombre,
            id__lt=visita.id,
            estado__in=["documentos_enviados", "en_revision_documentos", "confirmada"],
        ).order_by("-id")
        for visita_anterior in visitas_anteriores:
            asistentes_del_programa = visita_anterior.asistentes.filter(
                puede_reutilizar=True
            ).values_list(
                "id",
                "nombre_completo",
                "tipo_documento",
                "numero_documento",
                "correo",
                "telefono",
            )
            asistentes_previos.extend(asistentes_del_programa)
        # Eliminar duplicados basados en número de documento
        nums_doc_vistos = set()
        asistentes_previos_unicos = []
        for doc_info in asistentes_previos:
            num_doc = doc_info[3]  # numero_documento
            if num_doc not in nums_doc_vistos:
                nums_doc_vistos.add(num_doc)
                asistentes_previos_unicos.append(doc_info)
        asistentes_previos = asistentes_previos_unicos

    # Sincronizar documento de salud desde Aprendiz -> Asistente (solo visitas internas)
    if tipo == "interna":
        from panel_instructor_interno.models import Aprendiz

        doc_salud_default = Documento.objects.filter(
            categoria="Formato Auto Reporte Condiciones de Salud"
        ).first()

        for asistente in asistentes:
            tiene_doc_salud = asistente.documentos_subidos.filter(
                documento_requerido__categoria="Formato Auto Reporte Condiciones de Salud"
            ).exists()

            if not tiene_doc_salud:
                aprendiz = Aprendiz.objects.filter(
                    ficha__numero=visita.numero_ficha,
                    numero_documento=asistente.numero_documento,
                ).first()

                if aprendiz:
                    docs_salud_aprendiz = DocumentoSubidoAprendiz.objects.filter(
                        aprendiz=aprendiz,
                        documento_requerido__categoria="Formato Auto Reporte Condiciones de Salud",
                    ).select_related("documento_requerido")

                    for doc_subido in docs_salud_aprendiz:
                        DocumentoSubidoAsistente.objects.update_or_create(
                            documento_requerido=doc_subido.documento_requerido,
                            asistente_interna=asistente,
                            asistente_externa=None,
                            defaults={"archivo": doc_subido.archivo},
                        )

                    # Compatibilidad con registros antiguos que guardaban en documento_adicional
                    if (
                        not docs_salud_aprendiz.exists()
                        and aprendiz.documento_adicional
                        and doc_salud_default
                    ):
                        DocumentoSubidoAsistente.objects.update_or_create(
                            documento_requerido=doc_salud_default,
                            asistente_interna=asistente,
                            asistente_externa=None,
                            defaults={"archivo": aprendiz.documento_adicional},
                        )

            asistente.tiene_doc_salud = asistente.documentos_subidos.filter(
                documento_requerido__categoria="Formato Auto Reporte Condiciones de Salud"
            ).exists()
    else:
        for asistente in asistentes:
            asistente.tiene_doc_salud = asistente.documentos_subidos.filter(
                documento_requerido__categoria="Formato Auto Reporte Condiciones de Salud"
            ).exists()

    # Obtener documentos disponibles para descargar, agrupados por categoria
    documentos_disponibles = Documento.objects.all().order_by(
        "categoria", "-fecha_subida"
    )
    documentos_por_categoria = _agrupar_documentos_por_categoria(documentos_disponibles)
    documentos_registro_por_categoria = {
        categoria: docs
        for categoria, docs in documentos_por_categoria.items()
        if _es_categoria_documento_registro(categoria)
    }
    documentos_finales_por_categoria = {
        categoria: docs
        for categoria, docs in documentos_por_categoria.items()
        if _es_categoria_archivo_final(categoria)
    }

    asistentes_actuales = asistentes.count()
    solicitud_final_historica = _solicitud_final_ya_enviada(visita, tipo)
    puede_agregar = asistentes_actuales < max_asistentes and not solicitud_final_historica
    # Permitir archivos finales si hay al menos 1 asistente registrado
    mostrar_archivos_finales = asistentes_actuales > 0

    resumen_pendientes = _resumen_pendientes_correccion(visita, tipo)
    flujo_completo_con_finales = (
        asistentes_actuales >= max_asistentes
        and resumen_pendientes["archivos_finales_faltantes"] == 0
    )
    puede_actualizar_asistentes = not (
        solicitud_final_historica or flujo_completo_con_finales
    )

    asistentes_documentos_pendientes = []
    for asistente in asistentes:
        faltantes = []
        if not getattr(asistente, "tiene_doc_salud", False):
            faltantes.append("Reporte de condiciones de salud")

        if asistente.tipo_documento == "TI" and not getattr(
            asistente, "formato_autorizacion_padres", None
        ):
            faltantes.append("Autorización de padres")

        if faltantes:
            asistentes_documentos_pendientes.append(
                {
                    "id": asistente.id,
                    "nombre": asistente.nombre_completo,
                    "tipo_documento": asistente.tipo_documento,
                    "documento": f"{asistente.get_tipo_documento_display()} {asistente.numero_documento}",
                    "faltantes": faltantes,
                }
            )

    context = {
        "visita": visita,
        "tipo": tipo,
        "asistentes": asistentes,
        "asistentes_actuales": asistentes_actuales,
        "max_asistentes": max_asistentes,
        "puede_agregar": puede_agregar,
        "documentos_registro_por_categoria": documentos_registro_por_categoria,
        "documentos_finales_por_categoria": documentos_finales_por_categoria,
        "asistentes_previos": asistentes_previos,
        "tiene_asistentes_previos": len(asistentes_previos) > 0,
        "mostrar_archivos_finales": mostrar_archivos_finales,
        "puede_actualizar_asistentes": puede_actualizar_asistentes,
        "solicitud_final_historica": solicitud_final_historica,
        "asistentes_documentos_pendientes": asistentes_documentos_pendientes,
        "total_asistentes_pendientes": len(asistentes_documentos_pendientes),
    }

    if request.method == "POST" and request.POST.get("accion") == "carga_masiva_asistentes":
        if tipo != "externa":
            messages.error(
                request,
                "La carga masiva de asistentes solo está disponible para visitas externas.",
            )
            return redirect(
                "panel_visitante:registrar_asistentes",
                tipo=tipo,
                visita_id=visita_id,
            )

        if solicitud_final_historica:
            messages.error(
                request,
                "La solicitud final ya fue enviada previamente. No es posible modificar los datos de asistentes.",
            )
            return redirect(
                "panel_visitante:registrar_asistentes",
                tipo=tipo,
                visita_id=visita_id,
            )

        archivo = request.FILES.get("archivo_carga_masiva")
        if not archivo:
            messages.error(request, "Seleccione un archivo para realizar la carga masiva.")
            return redirect(
                "panel_visitante:registrar_asistentes",
                tipo=tipo,
                visita_id=visita_id,
            )

        filas, errores_lectura = _extraer_filas_importacion_asistentes(archivo)
        if errores_lectura:
            for error in errores_lectura:
                messages.error(request, error)
            return redirect(
                "panel_visitante:registrar_asistentes",
                tipo=tipo,
                visita_id=visita_id,
            )

        resumen_importacion = _procesar_carga_masiva_asistentes(
            visita=visita,
            tipo=tipo,
            filas=filas,
            max_asistentes=max_asistentes,
        )

        if resumen_importacion["creados"]:
            messages.success(
                request,
                f"Carga masiva completada. Se registraron {resumen_importacion['creados']} asistentes.",
            )

        if resumen_importacion["duplicados"]:
            messages.warning(
                request,
                f"Se omitieron {resumen_importacion['duplicados']} filas por documento repetido en esta visita.",
            )

        if resumen_importacion["omitidos_limite"]:
            messages.warning(
                request,
                f"Se omitieron {resumen_importacion['omitidos_limite']} filas por límite máximo de asistentes.",
            )

        for error in resumen_importacion["errores"][:10]:
            messages.error(request, error)

        if len(resumen_importacion["errores"]) > 10:
            messages.error(
                request,
                f"Se ocultaron {len(resumen_importacion['errores']) - 10} errores adicionales para mantener legible el resultado.",
            )

        return redirect(
            "panel_visitante:registrar_asistentes",
            tipo=tipo,
            visita_id=visita_id,
        )

    # Procesar archivos finales si se envía el formulario del modal
    if (
        request.method == "POST"
        and mostrar_archivos_finales  # Permitir si hay al menos 1 asistente
        and any(f.startswith("archivo_final_") for f in request.FILES)
    ):
        extensiones_permitidas = {".pdf", ".doc", ".docx"}
        archivos_subidos = []
        archivos_invalidos = []
        archivos_a_guardar = []
        # Obtener el primer asistente para asociar los archivos finales
        primer_asistente = asistentes.first() if asistentes.exists() else None

        if primer_asistente:
            for docs in context["documentos_finales_por_categoria"].values():
                for doc in docs:
                    archivo = request.FILES.get(f"archivo_final_{doc.id}")
                    if archivo:
                        # Guardar una nueva versión para conservar historial de envíos.
                        DocumentoSubidoAsistente.objects.create(
                            documento_requerido=doc,
                            asistente_interna=(
                                primer_asistente if tipo == "interna" else None
                            ),
                            asistente_externa=(
                                primer_asistente if tipo == "externa" else None
                            ),
                            archivo=archivo,
                            estado="pendiente",
                            observaciones_revision="",
                        )
                        archivos_subidos.append(doc.titulo)

            if archivos_subidos:
                _sincronizar_estado_asistente_por_docs(primer_asistente, tipo)

                if visita.estado in ["documentos_enviados", "en_revision_documentos"]:
                    visita.estado = "aprobada_inicial"
                    visita.save(update_fields=["estado"])

                resumen_pendientes = _resumen_pendientes_correccion(visita, tipo)
                auto_reenvio_aplicado = _auto_reenviar_a_revision_si_aplica(
                    visita, tipo, resumen_pendientes
                )

                if es_ajax:
                    return JsonResponse(
                        {
                            "success": True,
                            "message": "Archivos finales corregidos y cargados correctamente.",
                            "archivos_subidos": archivos_subidos,
                            **resumen_pendientes,
                            "auto_reenvio_aplicado": auto_reenvio_aplicado,
                            "listo_para_confirmar_envio": resumen_pendientes[
                                "pendientes_correccion"
                            ]
                            == 0,
                        }
                    )

                messages.success(request, "Archivos finales subidos con éxito.")
                return redirect(
                    "panel_visitante:registrar_asistentes",
                    tipo=tipo,
                    visita_id=visita_id,
                )
            else:
                if es_ajax:
                    return JsonResponse(
                        {
                            "success": False,
                            "error": "No se subieron archivos finales.",
                        },
                        status=400,
                    )
                messages.warning(request, "No se subieron archivos finales.")
        else:
            if es_ajax:
                return JsonResponse(
                    {
                        "success": False,
                        "error": "No hay asistentes registrados para asociar los archivos.",
                    },
                    status=400,
                )
            messages.error(
                request, "No hay asistentes registrados para asociar los archivos."
            )

    if request.method == "POST":
        if solicitud_final_historica:
            messages.error(
                request,
                "La solicitud final ya fue enviada previamente. No es posible modificar los datos de asistentes.",
            )
        elif not puede_agregar:
            messages.error(
                request, f"Ya se alcanzó el límite de {max_asistentes} asistentes."
            )
        else:
            nombre = request.POST.get("nombre_completo", "").strip()
            tipo_doc = request.POST.get("tipo_documento", "")
            num_doc = request.POST.get("numero_documento", "").strip()
            correo_asistente = request.POST.get("correo", "").strip().lower()
            telefono = request.POST.get("telefono", "").strip()

            # Validar campos obligatorios y formato
            campos_ok = True
            errores_campos = []
            tipos_documento_validos = {"CC", "CE", "TI", "PPT", "PP"}

            nombre = " ".join(nombre.split())
            if not nombre or not tipo_doc or not num_doc:
                errores_campos.append("Complete todos los campos obligatorios.")

            if nombre:
                if len(nombre) < 3 or len(nombre) > 80:
                    errores_campos.append(
                        "El nombre completo debe tener entre 3 y 80 caracteres."
                    )
                elif not re.fullmatch(r"[A-Za-zÁÉÍÓÚáéíóúÑñÜü\s]+", nombre):
                    errores_campos.append(
                        "El nombre completo solo puede contener letras y espacios."
                    )

            if tipo_doc and tipo_doc not in tipos_documento_validos:
                errores_campos.append("Seleccione un tipo de documento valido.")

            if num_doc:
                if not num_doc.isdigit():
                    errores_campos.append(
                        "El numero de documento solo debe contener numeros."
                    )
                elif not 5 <= len(num_doc) <= 10:
                    errores_campos.append(
                        "El numero de documento debe tener entre 5 y 10 digitos."
                    )

            if not correo_asistente:
                errores_campos.append("El correo electronico es obligatorio.")
            else:
                try:
                    validate_email(correo_asistente)
                except ValidationError:
                    errores_campos.append("Ingrese un correo electronico valido.")

            if telefono:
                if not telefono.isdigit():
                    errores_campos.append("El telefono solo debe contener numeros.")
                elif not 7 <= len(telefono) <= 10:
                    errores_campos.append(
                        "El telefono debe tener entre 7 y 10 digitos."
                    )

            if tipo == "interna":
                documento_existente = AsistenteVisitaInterna.objects.filter(
                    visita=visita, numero_documento=num_doc
                ).exists()
            else:
                documento_existente = AsistenteVisitaExterna.objects.filter(
                    visita=visita, numero_documento=num_doc
                ).exists()

            if num_doc and documento_existente:
                errores_campos.append(
                    "Este numero de documento ya esta registrado para esta visita."
                )

            if errores_campos:
                campos_ok = False
                for error in errores_campos:
                    messages.error(request, error)

            documentos_disponibles = Documento.objects.all().order_by(
                "categoria", "-fecha_subida"
            )
            documentos_por_categoria = _agrupar_documentos_por_categoria(
                documentos_disponibles
            )
            documentos_registro_por_categoria = {
                categoria: docs
                for categoria, docs in documentos_por_categoria.items()
                if _es_categoria_documento_registro(categoria)
            }

            # Validar que solo el documento de 'Formato Auto Reporte Condiciones de Salud' fue subido
            archivos_ok = False
            archivos_dict = {}
            extensiones_permitidas_docs = {".pdf", ".doc", ".docx"}
            archivos_invalidos = []
            for categoria, docs in documentos_registro_por_categoria.items():
                if categoria == "Formato Auto Reporte Condiciones de Salud":
                    for doc in docs:
                        file_field = f"documento_{doc.id}"
                        archivo = request.FILES.get(file_field)
                        if archivo:
                            extension = Path(archivo.name).suffix.lower()
                            if extension not in extensiones_permitidas_docs:
                                archivos_invalidos.append(doc.categoria or archivo.name)
                            else:
                                archivos_ok = True
                        archivos_dict[doc.id] = archivo
                elif categoria == "Formato Autorización Padres de Familia":
                    for doc in docs:
                        file_field = f"documento_{doc.id}"
                        archivo = request.FILES.get(file_field)
                        if archivo:
                            extension = Path(archivo.name).suffix.lower()
                            if extension not in extensiones_permitidas_docs:
                                archivos_invalidos.append(doc.categoria or archivo.name)
                        # Este archivo es opcional, por lo que no afecta archivos_ok
                        archivos_dict[doc.id] = archivo

            if archivos_invalidos:
                docs_invalidos = ", ".join(archivos_invalidos)
                messages.error(
                    request,
                    f"Solo se permiten archivos PDF o Word (.doc, .docx). Revise: {docs_invalidos}.",
                )
                campos_ok = False

            if tipo_doc == "TI":
                tiene_autorizacion_ti = any(
                    archivo
                    for categoria, docs in documentos_registro_por_categoria.items()
                    if categoria == "Formato Autorización Padres de Familia"
                    for doc in docs
                    for archivo in [archivos_dict.get(doc.id)]
                )
                if not tiene_autorizacion_ti:
                    messages.error(
                        request,
                        "Para Tarjeta de Identidad (TI), la autorización de padres es obligatoria.",
                    )
                    campos_ok = False

            if campos_ok and archivos_ok:
                try:
                    if tipo == "interna":
                        asistente = AsistenteVisitaInterna.objects.create(
                            visita=visita,
                            nombre_completo=nombre,
                            tipo_documento=tipo_doc,
                            numero_documento=num_doc,
                            correo=correo_asistente,
                            telefono=telefono,
                        )
                    else:
                        asistente = AsistenteVisitaExterna.objects.create(
                            visita=visita,
                            nombre_completo=nombre,
                            tipo_documento=tipo_doc,
                            numero_documento=num_doc,
                            correo=correo_asistente,
                            telefono=telefono,
                        )
                    # Guardar archivos subidos

                    # Separar el formato de autorización de padres de los demás documentos
                    formato_padres_archivo = None
                    documentos_regulares = {}

                    for doc_id, archivo in archivos_dict.items():
                        if archivo:
                            # Verificar si es el documento de autorización de padres
                            documento = Documento.objects.get(id=doc_id)
                            if (
                                documento.categoria
                                == "Formato Autorización Padres de Familia"
                            ):
                                formato_padres_archivo = archivo
                            else:
                                documentos_regulares[doc_id] = archivo

                    # Guardar el formato de autorización de padres directamente en el asistente
                    if formato_padres_archivo:
                        asistente.formato_autorizacion_padres = formato_padres_archivo
                        asistente.save()

                    # Guardar los demás documentos en DocumentoSubidoAsistente
                    for doc_id, archivo in documentos_regulares.items():
                        DocumentoSubidoAsistente.objects.create(
                            documento_requerido_id=doc_id,
                            asistente_interna=(
                                asistente if tipo == "interna" else None
                            ),
                            asistente_externa=(
                                asistente if tipo == "externa" else None
                            ),
                            archivo=archivo,
                        )

                    # Si es visita interna, también registrar aprendiz en la ficha
                    if tipo == "interna":
                        try:
                            from panel_instructor_interno.models import Ficha, Aprendiz

                            # Obtener la ficha de la visita
                            ficha = Ficha.objects.get(numero=visita.numero_ficha)

                            # Verificar si el aprendiz ya existe en la ficha
                            aprendiz_existe = Aprendiz.objects.filter(
                                ficha=ficha, numero_documento=num_doc
                            ).exists()

                            if not aprendiz_existe:
                                # Crear aprendiz en la ficha con el documento de salud
                                aprendiz_data = {
                                    "ficha": ficha,
                                    "nombre": (
                                        nombre.split()[0] if nombre else ""
                                    ),  # Primer nombre
                                    "apellido": (
                                        " ".join(nombre.split()[1:])
                                        if len(nombre.split()) > 1
                                        else ""
                                    ),  # Resto como apellido
                                    "tipo_documento": tipo_doc,
                                    "numero_documento": num_doc,
                                    "correo": correo_asistente,
                                    "telefono": telefono,
                                    "estado": "activo",
                                }

                                # Si hay documento de salud, asignarlo
                                for doc_id, archivo in archivos_dict.items():
                                    if archivo and doc_id in archivos_dict:
                                        # El archivo del Auto Reporte va a documento_adicional
                                        aprendiz_data["documento_adicional"] = archivo

                                Aprendiz.objects.create(**aprendiz_data)
                        except Exception as e:
                            # Si falla la creación del aprendiz en ficha, no interrumpir el flujo
                            pass

                    # El QR se envia unicamente cuando la visita queda confirmada en su totalidad.
                    messages.success(
                        request,
                        f'Asistente "{nombre}" registrado correctamente. El QR se enviara cuando la visita sea confirmada definitivamente.',
                    )

                    # Ya no es necesario guardar en session, el context lo calcula dinámicamente
                    return redirect(
                        "panel_visitante:registrar_asistentes",
                        tipo=tipo,
                        visita_id=visita_id,
                    )
                except Exception as e:
                    if "unique" in str(e).lower():
                        messages.error(
                            request,
                            "Este documento ya está registrado para esta visita.",
                        )
                    else:
                        messages.error(request, f"Error al registrar: {str(e)}")
            else:
                if campos_ok and not archivos_ok:
                    messages.error(request, "Debe subir todos los archivos requeridos.")

    # Obtener documentos disponibles para descargar, agrupados por categoría
    documentos_disponibles = Documento.objects.all().order_by(
        "categoria", "-fecha_subida"
    )
    documentos_por_categoria = _agrupar_documentos_por_categoria(documentos_disponibles)
    documentos_registro_por_categoria = {
        categoria: docs
        for categoria, docs in documentos_por_categoria.items()
        if _es_categoria_documento_registro(categoria)
    }
    documentos_finales_por_categoria = {
        categoria: docs
        for categoria, docs in documentos_por_categoria.items()
        if _es_categoria_archivo_final(categoria)
    }

    context = {
        "visita": visita,
        "tipo": tipo,
        "asistentes": asistentes,
        "asistentes_actuales": asistentes_actuales,
        "max_asistentes": max_asistentes,
        "puede_agregar": puede_agregar,
        "documentos_registro_por_categoria": documentos_registro_por_categoria,
        "documentos_finales_por_categoria": documentos_finales_por_categoria,
        "asistentes_previos": asistentes_previos,
        "tiene_asistentes_previos": len(asistentes_previos) > 0,
        "mostrar_archivos_finales": mostrar_archivos_finales,
        "puede_actualizar_asistentes": puede_actualizar_asistentes,
        "solicitud_final_historica": solicitud_final_historica,
        "asistentes_documentos_pendientes": asistentes_documentos_pendientes,
        "total_asistentes_pendientes": len(asistentes_documentos_pendientes),
    }

    return render(request, "registrar_asistentes.html", context)


def eliminar_asistente(request, tipo, asistente_id):
    """
    Eliminar un asistente de una visita.
    """
    # Verificar autenticación
    if not request.session.get("responsable_autenticado"):
        return redirect("panel_visitante:login_responsable")

    correo = request.session.get("responsable_correo")
    if tipo == "interna":
        asistente = get_object_or_404(AsistenteVisitaInterna, id=asistente_id)
        visita = asistente.visita
        # Verificar que el responsable tenga acceso a esta visita
        if not _tiene_acceso_por_correo(visita, correo):
            messages.error(request, "No tiene permiso para esta acción.")
            return _redirect_segun_rol(request)
        if _solicitud_final_ya_enviada(visita, tipo):
            messages.error(
                request,
                "La solicitud final ya fue enviada previamente. No es posible eliminar asistentes.",
            )
            return redirect(
                "panel_visitante:registrar_asistentes",
                tipo=tipo,
                visita_id=visita.id,
            )
        visita_id = visita.id
        asistente.delete()
    elif tipo == "externa":
        asistente = get_object_or_404(AsistenteVisitaExterna, id=asistente_id)
        visita = asistente.visita
        if not _tiene_acceso_por_correo(visita, correo):
            messages.error(request, "No tiene permiso para esta acción.")
            return _redirect_segun_rol(request)
        if visita.estado in ["documentos_enviados", "en_revision_documentos", "confirmada"]:
            messages.error(
                request,
                "No es posible eliminar asistentes en el estado actual de la visita.",
            )
            return redirect(
                "panel_visitante:registrar_asistentes",
                tipo=tipo,
                visita_id=visita.id,
            )
        visita_id = visita.id
        asistente.delete()
    else:
        messages.error(request, "Tipo de visita no válido.")
        return _redirect_segun_rol(request)

    messages.success(request, "Asistente eliminado correctamente.")
    return redirect(
        "panel_visitante:registrar_asistentes", tipo=tipo, visita_id=visita_id
    )


def enviar_solicitud_final(request, tipo, visita_id):
    """
    Vista para que el responsable envíe la solicitud final de aprobación.
    Cambia el estado de la visita a 'documentos_enviados' para revisión de documentos.
    """
    # Verificar autenticación
    es_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"

    def respuesta_error(mensaje, status_code=400):
        if es_ajax:
            return JsonResponse(
                {
                    "success": False,
                    "error": mensaje,
                },
                status=status_code,
            )
        messages.error(request, mensaje)
        return _redirect_segun_rol(request, tipo, visita_id)

    def respuesta_ok(mensaje):
        if es_ajax:
            return JsonResponse(
                {
                    "success": True,
                    "message": mensaje,
                }
            )
        messages.success(request, mensaje)
        return _redirect_segun_rol(request, tipo, visita_id)

    if not request.session.get("responsable_autenticado"):
        if es_ajax:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Sesión no válida. Inicie sesión nuevamente.",
                },
                status=401,
            )
        return redirect("panel_visitante:login_responsable")

    correo = request.session.get("responsable_correo")
    # Obtener la visita
    if tipo == "interna":
        visita = get_object_or_404(VisitaInterna, id=visita_id)
        if not _tiene_acceso_por_correo(visita, correo):
            if es_ajax:
                return JsonResponse(
                    {
                        "success": False,
                        "error": "No tiene permiso para esta acción.",
                    },
                    status=403,
                )
            messages.error(request, "No tiene permiso para esta acción.")
            return _redirect_segun_rol(request)
    elif tipo == "externa":
        visita = get_object_or_404(VisitaExterna, id=visita_id)
        if not _tiene_acceso_por_correo(visita, correo):
            if es_ajax:
                return JsonResponse(
                    {
                        "success": False,
                        "error": "No tiene permiso para esta acción.",
                    },
                    status=403,
                )
            messages.error(request, "No tiene permiso para esta acción.")
            return _redirect_segun_rol(request)
    else:
        if es_ajax:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Tipo de visita no válido.",
                },
                status=400,
            )
        messages.error(request, "Tipo de visita no válido.")
        return _redirect_segun_rol(request)

    # Hacer el endpoint idempotente cuando ya esta en revision o confirmada.
    if visita.estado in ["documentos_enviados", "en_revision_documentos", "confirmada"]:
        return respuesta_ok(
            "La solicitud ya fue enviada y la visita se encuentra en proceso de revisión.",
        )

    # Verificar que la visita esté en estado aprobada_inicial
    if visita.estado != "aprobada_inicial":
        return respuesta_error(
            "Solo puede enviar la solicitud final cuando la visita esté aprobada inicialmente.",
        )

    cantidad_maxima = (
        visita.cantidad_aprendices if tipo == "interna" else visita.cantidad_visitantes
    )
    if cantidad_maxima < 1:
        return respuesta_error(
            "La visita debe tener una cantidad de asistentes mayor a cero antes de enviar la solicitud final.",
        )

    # Verificar que haya al menos un asistente registrado
    if visita.asistentes.count() == 0:
        return respuesta_error(
            "Debe registrar al menos un asistente antes de enviar la solicitud final.",
        )

    # Evitar reenvío mientras existan rechazos pendientes de corrección
    if visita.asistentes.filter(estado="documentos_rechazados").exists():
        return respuesta_error(
            "Hay asistentes con documentos rechazados. Corrija los archivos antes de reenviar la solicitud.",
        )

    # Cambiar el estado a documentos_enviados
    visita.estado = "documentos_enviados"
    visita.save()

    # Guardar una marca historica para bloquear futuras ediciones de asistentes,
    # incluso si la visita vuelve temporalmente a aprobada_inicial por correcciones.
    try:
        usuario_historial = (
            request.user
            if getattr(request, "user", None)
            and getattr(request.user, "is_authenticated", False)
            else None
        )
        descripcion_historial = (
            f"{MARCADOR_SOLICITUD_FINAL_ENVIADA} Envio final realizado por el responsable {correo}."
        )

        if tipo == "interna":
            HistorialAccionVisitaInterna.objects.create(
                visita=visita,
                usuario=usuario_historial,
                tipo_accion="modificacion",
                descripcion=descripcion_historial,
                ip_address=request.META.get("REMOTE_ADDR"),
            )
        else:
            HistorialAccionVisitaExterna.objects.create(
                visita=visita,
                usuario=usuario_historial,
                tipo_accion="modificacion",
                descripcion=descripcion_historial,
                ip_address=request.META.get("REMOTE_ADDR"),
            )
    except Exception:
        pass

    # Enviar correo al responsable informando que la solicitud final fue enviada
    try:
        if tipo == "interna":
            panel_path = reverse("panel_instructor_interno:panel")
            template_name = "emails/solicitud_final_enviada_interna.html"
        else:
            panel_path = reverse("panel_instructor_externo:panel")
            template_name = "emails/solicitud_final_enviada_externa.html"
        panel_url = request.build_absolute_uri(panel_path)
        subject = "Solicitud final enviada con éxito — documentos en revisión"
        context = {
            "responsable_nombre": request.session.get("responsable_nombre", ""),
            "panel_url": panel_url,
        }
        html_content = render_to_string(template_name, context)
        text_content = strip_tags(html_content)
        msg = EmailMultiAlternatives(
            subject,
            text_content,
            settings.DEFAULT_FROM_EMAIL,
            [visita.correo_responsable],
        )
        msg.attach_alternative(html_content, "text/html")
        msg.send(fail_silently=True)
    except Exception:
        pass

    return respuesta_ok(
        "¡Solicitud final enviada correctamente! El administrador revisará los documentos de los asistentes.",
    )


def restablecer_contraseña(request):
    """
    Vista para mostrar la página de restablecimiento de contraseña.
    Maneja la solicitud de recuperación y el cambio de contraseña.
    """
    if request.method == "POST":
        form = PasswordResetRequestForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data["email"]

            # Buscar visitante con ese email
            visitante = RegistroVisitante.objects.filter(correo__iexact=email).first()

            if not visitante:
                form.add_error("email", "Este correo no está registrado.")
            else:
                # Generar token y uid usando Django's default_token_generator
                token = default_token_generator.make_token(visitante)
                uid = urlsafe_base64_encode(force_bytes(visitante.pk))

                # Construir URL de reset usando reverse
                reset_path = reverse(
                    "panel_visitante:restablecer_contraseña_confirm",
                    kwargs={"uidb64": uid, "token": token},
                )
                reset_url = request.build_absolute_uri(reset_path)

                # Preparar contexto para el template HTML
                email_context = {
                    "nombre": visitante.documento,
                    "reset_url": reset_url,
                    "year": datetime.now().year,
                }

                # Renderizar template HTML
                html_content = render_to_string(
                    "email_restablecer_visitante.html", email_context
                )
                text_content = strip_tags(html_content)

                # Enviar correo HTML
                subject = "🔐 Recuperación de Contraseña - MINEGATE"
                email_msg = EmailMultiAlternatives(
                    subject=subject,
                    body=text_content,
                    from_email=settings.DEFAULT_FROM_EMAIL or "noreply@minegate.com",
                    to=[visitante.correo],
                )
                email_msg.attach_alternative(html_content, "text/html")
                try:
                    email_msg.send()
                    messages.success(
                        request,
                        "Se ha enviado un correo con instrucciones para restablecer tu contraseña.",
                    )
                    return redirect("panel_visitante:correo_enviado")
                except Exception as e:
                    messages.warning(
                        request,
                        "El correo no se pudo enviar. Por favor intenta más tarde.",
                    )
                    print(f"Error enviando email: {str(e)}")
    else:
        form = PasswordResetRequestForm()

    context = {"form": form, "titulo": "Recuperar Contraseña"}
    return render(request, "solicitar_recuperacion_visitante.html", context)


def correo_enviado_view(request):
    """
    Vista que se muestra después de enviar el correo de recuperación
    """
    return render(request, "correo_enviado_visitante.html")


@csrf_protect
def restablecer_contraseña_confirm(request, uidb64, token):
    """
    Vista para confirmar y restablecer la contraseña con token
    """
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        visitante = RegistroVisitante.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, RegistroVisitante.DoesNotExist):
        visitante = None

    if visitante is not None and default_token_generator.check_token(visitante, token):
        if request.method == "POST":
            form = PasswordResetConfirmForm(request.POST)
            if form.is_valid():
                password = form.cleaned_data["password1"]
                visitante.set_password(password)
                visitante.save()

                messages.success(
                    request,
                    "¡Tu contraseña ha sido actualizada exitosamente! Ahora puedes iniciar sesión.",
                )
                return redirect("panel_visitante:contraseña_actualizada")
        else:
            form = PasswordResetConfirmForm()

        context = {
            "form": form,
            "validlink": True,
            "titulo": "Establecer Nueva Contraseña",
        }
        return render(request, "restablecer_contraseña_confirm_visitante.html", context)
    else:
        messages.error(request, "El enlace de recuperación es inválido o ha expirado.")
        context = {"validlink": False, "titulo": "Enlace Inválido"}
        return render(request, "restablecer_contraseña_confirm_visitante.html", context)


def contraseña_actualizada_view(request):
    """
    Vista que se muestra después de actualizar la contraseña
    """
    return render(request, "contraseña_actualizada_visitante.html")


def actualizar_perfil(request):
    """
    Vista para que el usuario actualice sus datos personales y contraseña
    """
    embed_mode = request.GET.get("embed") == "1" or request.POST.get("embed") == "1"

    # Verificar que el usuario esté autenticado
    if not request.session.get("responsable_autenticado"):
        messages.warning(
            request,
            "Debe iniciar sesión para acceder a esta página.",
            extra_tags=AUTH_VISITANTE_MESSAGE_TAG,
        )
        return redirect("panel_visitante:login_responsable")

    documento = request.session.get("responsable_documento")
    rol = request.session.get("responsable_rol")
    visitante = RegistroVisitante.objects.filter(documento=documento).first()

    if not visitante:
        messages.error(
            request,
            "No se encontró el usuario.",
            extra_tags=AUTH_VISITANTE_MESSAGE_TAG,
        )
        return redirect("panel_visitante:login_responsable")

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "actualizar_datos":
            form_perfil = ActualizarPerfilForm(request.POST, current_user=visitante)
            form_contrasena = CambiarContrasenaForm()

            if form_perfil.is_valid():
                visitante.nombre = form_perfil.cleaned_data["nombre"]
                visitante.apellido = form_perfil.cleaned_data["apellido"]
                visitante.tipo_documento = form_perfil.cleaned_data["tipo_documento"]
                visitante.telefono = form_perfil.cleaned_data["telefono"]
                visitante.correo = form_perfil.cleaned_data["correo"]
                visitante.save()

                # Actualizar sesión
                request.session["responsable_nombre"] = visitante.nombre
                request.session["responsable_apellido"] = visitante.apellido
                request.session["responsable_tipo_documento"] = visitante.tipo_documento
                request.session["responsable_telefono"] = visitante.telefono
                request.session["responsable_correo"] = visitante.correo
                request.session.modified = True

                messages.success(
                    request, "✅ Tus datos han sido actualizados exitosamente."
                )
                return redirect("panel_visitante:actualizar_perfil")

        elif action == "cambiar_contrasena":
            form_perfil = ActualizarPerfilForm(
                initial={
                    "nombre": visitante.nombre,
                    "apellido": visitante.apellido,
                    "tipo_documento": visitante.tipo_documento,
                    "telefono": visitante.telefono,
                    "correo": visitante.correo,
                },
                current_user=visitante,
            )
            form_contrasena = CambiarContrasenaForm(request.POST)

            if form_contrasena.is_valid():
                contrasena_actual = form_contrasena.cleaned_data["contrasena_actual"]

                if visitante.check_password(contrasena_actual):
                    nueva_contrasena = form_contrasena.cleaned_data["nueva_contrasena"]

                    if visitante.check_password(nueva_contrasena):
                        messages.error(
                            request,
                            "La nueva contraseña no puede ser igual a la actual.",
                        )
                        return redirect("panel_visitante:actualizar_perfil")

                    visitante.set_password(nueva_contrasena)
                    visitante.save()

                    messages.success(
                        request, "🔒 Tu contraseña ha sido actualizada exitosamente."
                    )
                    return redirect("panel_visitante:actualizar_perfil")
                else:
                    messages.error(request, "La contraseña actual es incorrecta.")
    else:
        form_perfil = ActualizarPerfilForm(
            initial={
                "nombre": visitante.nombre,
                "apellido": visitante.apellido,
                "tipo_documento": visitante.tipo_documento,
                "telefono": visitante.telefono,
                "correo": visitante.correo,
            },
            current_user=visitante,
        )
        form_contrasena = CambiarContrasenaForm()

    context = {
        "form_perfil": form_perfil,
        "form_contrasena": form_contrasena,
        "visitante": visitante,
        "titulo": "Actualizar Perfil",
        "embed_mode": embed_mode,
        "volver_url": (
            reverse("panel_instructor_interno:panel")
            if rol == "interno"
            else (
                reverse("panel_instructor_externo:panel")
                if rol == "externo"
                else reverse("panel_visitante:panel_responsable")
            )
        ),
    }
    return render(request, "actualizar_perfil.html", context)


def actualizar_documento_asistente(request, tipo, asistente_id):
    """
    Permite actualizar documentos de un asistente.
    Soporta respuesta JSON cuando la solicitud es AJAX.
    """
    es_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"

    if not request.session.get("responsable_autenticado"):
        if es_ajax:
            return JsonResponse(
                {"success": False, "error": "No autenticado."}, status=401
            )
        return redirect("panel_visitante:login_responsable")

    correo = request.session.get("responsable_correo")

    if tipo == "interna":
        asistente = get_object_or_404(AsistenteVisitaInterna, id=asistente_id)
        visita = asistente.visita
    elif tipo == "externa":
        asistente = get_object_or_404(AsistenteVisitaExterna, id=asistente_id)
        visita = asistente.visita
    else:
        error_msg = "Tipo de visita no válido."
        if es_ajax:
            return JsonResponse({"success": False, "error": error_msg}, status=400)
        messages.error(request, error_msg)
        return _redirect_segun_rol(request)

    if not _tiene_acceso_por_correo(visita, correo):
        error_msg = "No tiene permiso para esta acción."
        if es_ajax:
            return JsonResponse({"success": False, "error": error_msg}, status=403)
        messages.error(request, error_msg)
        return _redirect_segun_rol(request)

    max_asistentes = (
        visita.cantidad_aprendices if tipo == "interna" else visita.cantidad_visitantes
    )
    asistentes_actuales = visita.asistentes.count()
    resumen_pendientes = _resumen_pendientes_correccion(visita, tipo)
    sin_pendientes_correccion = resumen_pendientes["pendientes_correccion"] == 0
    flujo_completo_con_finales = (
        asistentes_actuales >= max_asistentes
        and resumen_pendientes["archivos_finales_faltantes"] == 0
    )
    visita_confirmada = visita.estado == "confirmada"

    if visita_confirmada or (flujo_completo_con_finales and sin_pendientes_correccion):
        error_msg = (
            "Ya no se puede actualizar este asistente porque la visita está "
            "confirmada o no tiene correcciones pendientes."
        )
        if es_ajax:
            return JsonResponse({"success": False, "error": error_msg}, status=400)
        messages.warning(request, error_msg)
        return redirect(
            "panel_visitante:registrar_asistentes", tipo=tipo, visita_id=visita.id
        )

    if request.method != "POST":
        if es_ajax:
            return JsonResponse(
                {"success": False, "error": "Método no permitido."}, status=405
            )
        return redirect(
            "panel_visitante:registrar_asistentes", tipo=tipo, visita_id=visita.id
        )

    from documentos.models import Documento as DocumentoModel

    # Para AJAX llega "archivo_correccion". Para formulario tradicional,
    # se mantiene la compatibilidad con los nombres por asistente.
    documento_subido_id = (request.POST.get("documento_subido_id") or "").strip()
    tipo_correccion = (request.POST.get("tipo_correccion") or "documento").strip().lower()
    if tipo_correccion not in {"documento", "autorizacion_padres"}:
        tipo_correccion = "documento"

    archivo_correccion = request.FILES.get("archivo_correccion")
    archivo_autorizacion_correccion = request.FILES.get(
        "archivo_autorizacion_correccion"
    )
    if tipo_correccion == "autorizacion_padres":
        archivo_salud = request.FILES.get(f"documento_salud_{asistente_id}")
        archivo_autorizacion = (
            archivo_autorizacion_correccion
            or archivo_correccion
            or request.FILES.get(
            f"formato_padres_{asistente_id}"
            )
        )
    else:
        archivo_salud = (
            archivo_correccion
            or request.FILES.get("archivo_salud_correccion")
            or request.FILES.get(f"documento_salud_{asistente_id}")
        )
        archivo_autorizacion = (
            archivo_autorizacion_correccion
            or request.FILES.get(f"formato_padres_{asistente_id}")
        )

    if not archivo_salud and not archivo_autorizacion:
        error_msg = "Debe seleccionar al menos un archivo para corregir."
        if es_ajax:
            return JsonResponse({"success": False, "error": error_msg}, status=400)
        messages.error(request, error_msg)
        return redirect(
            "panel_visitante:registrar_asistentes", tipo=tipo, visita_id=visita.id
        )

    actualizaciones_asistente = {
        "estado": "pendiente_documentos",
        "observaciones_revision": "",
    }

    if archivo_salud:
        extensiones_permitidas_docs = {
            ".pdf",
            ".doc",
            ".docx",
            ".jpg",
            ".jpeg",
            ".png",
            ".xls",
            ".xlsx",
            ".ppt",
            ".pptx",
            ".txt",
        }
        extension = Path(archivo_salud.name).suffix.lower()
        if extension not in extensiones_permitidas_docs:
            error_msg = (
                "Formato no permitido para la corrección. "
                "Use PDF, Word, imagen, Excel, PowerPoint o TXT."
            )
            if es_ajax:
                return JsonResponse({"success": False, "error": error_msg}, status=400)
            messages.error(request, error_msg)
            return redirect(
                "panel_visitante:registrar_asistentes",
                tipo=tipo,
                visita_id=visita.id,
            )

        if documento_subido_id:
            filtros_doc = {"id": documento_subido_id}
            if tipo == "interna":
                filtros_doc["asistente_interna"] = asistente
            else:
                filtros_doc["asistente_externa"] = asistente

            doc_objetivo = (
                DocumentoSubidoAsistente.objects.select_related("documento_requerido")
                .filter(**filtros_doc)
                .first()
            )

            if not doc_objetivo:
                error_msg = "No se encontró el documento rechazado para este asistente."
                if es_ajax:
                    return JsonResponse(
                        {"success": False, "error": error_msg}, status=400
                    )
                messages.error(request, error_msg)
                return redirect(
                    "panel_visitante:registrar_asistentes",
                    tipo=tipo,
                    visita_id=visita.id,
                )

            documento_objetivo = doc_objetivo.documento_requerido
        else:
            documento_objetivo = DocumentoModel.objects.filter(
                categoria="Formato Auto Reporte Condiciones de Salud"
            ).first()
            if not documento_objetivo:
                error_msg = "No se encontró el documento base de salud para registrar la corrección."
                if es_ajax:
                    return JsonResponse(
                        {"success": False, "error": error_msg}, status=400
                    )
                messages.error(request, error_msg)
                return redirect(
                    "panel_visitante:registrar_asistentes",
                    tipo=tipo,
                    visita_id=visita.id,
                )

        DocumentoSubidoAsistente.objects.create(
            documento_requerido=documento_objetivo,
            asistente_interna=asistente if tipo == "interna" else None,
            asistente_externa=asistente if tipo == "externa" else None,
            archivo=archivo_salud,
            estado="pendiente",
            observaciones_revision="",
        )

    if archivo_autorizacion:
        actualizaciones_asistente.update(
            {
                "formato_autorizacion_padres": archivo_autorizacion,
                "estado_autorizacion_padres": "pendiente",
                "observaciones_autorizacion_padres": "",
            }
        )

    for campo, valor in actualizaciones_asistente.items():
        setattr(asistente, campo, valor)
    asistente.save(update_fields=list(actualizaciones_asistente.keys()))

    if visita.estado in ["documentos_enviados", "en_revision_documentos"]:
        visita.estado = "aprobada_inicial"
        visita.save(update_fields=["estado"])

    resumen_pendientes = _resumen_pendientes_correccion(visita, tipo)
    auto_reenvio_aplicado = _auto_reenviar_a_revision_si_aplica(
        visita, tipo, resumen_pendientes
    )

    success_msg = (
        f"Se actualizaron los archivos de {asistente.nombre_completo}. "
        "Ya puede reenviar la solicitud final."
    )
    if es_ajax:
        return JsonResponse(
            {
                "success": True,
                "message": success_msg,
                **resumen_pendientes,
                "auto_reenvio_aplicado": auto_reenvio_aplicado,
                "listo_para_confirmar_envio": resumen_pendientes[
                    "pendientes_correccion"
                ]
                == 0,
            }
        )

    messages.success(request, success_msg)
    return redirect(
        "panel_visitante:registrar_asistentes", tipo=tipo, visita_id=visita.id
    )


def actualizar_info_asistente(request, tipo, asistente_id):
    """Actualiza informacion basica del asistente y opcionalmente sus documentos."""
    embed_mode = request.GET.get("embed") == "1" or request.POST.get("embed") == "1"

    if not request.session.get("responsable_autenticado"):
        return redirect("panel_visitante:login_responsable")

    correo = request.session.get("responsable_correo")
    if tipo == "interna":
        asistente = get_object_or_404(AsistenteVisitaInterna, id=asistente_id)
        visita = asistente.visita
        if not _tiene_acceso_por_correo(visita, correo):
            messages.error(request, "No tiene permiso para esta accion.")
            return _redirect_segun_rol(request)
    elif tipo == "externa":
        asistente = get_object_or_404(AsistenteVisitaExterna, id=asistente_id)
        visita = asistente.visita
        if not _tiene_acceso_por_correo(visita, correo):
            messages.error(request, "No tiene permiso para esta accion.")
            return _redirect_segun_rol(request)
    else:
        messages.error(request, "Tipo de visita no valido.")
        return _redirect_segun_rol(request)

    def _contexto(extra=None):
        ctx = {
            "asistente": asistente,
            "tipo": tipo,
            "visita": visita,
            "tipo_documento_choices": asistente.TIPO_DOCUMENTO_CHOICES,
            "embed_mode": embed_mode,
        }
        if extra:
            ctx.update(extra)
        return ctx

    bloqueo_edicion = _solicitud_final_ya_enviada(visita, tipo)
    if tipo == "externa":
        bloqueo_edicion = visita.estado in [
            "documentos_enviados",
            "en_revision_documentos",
            "confirmada",
        ]

    if bloqueo_edicion:
        mensaje_bloqueo = (
            "La solicitud final ya fue enviada previamente. "
            "No puede modificar los datos del asistente."
        )
        if embed_mode:
            return render(
                request,
                "actualizar_info_asistente.html",
                _contexto(
                    {
                        "bloqueado_envio_final": True,
                        "mensaje_bloqueo": mensaje_bloqueo,
                    }
                ),
            )
        messages.warning(request, mensaje_bloqueo)
        return _redirect_segun_rol(request, tipo=tipo, visita_id=visita.id)

    if request.method == "POST":
        nombre = request.POST.get("nombre_completo", "").strip()
        tipo_doc = request.POST.get("tipo_documento", "").strip()
        numero_doc = request.POST.get("numero_documento", "").strip()
        correo_asistente = request.POST.get("correo", "").strip().lower()
        telefono = request.POST.get("telefono", "").strip()
        numero_doc_anterior = asistente.numero_documento
        archivo_salud = request.FILES.get("archivo_salud")
        archivo_autorizacion = request.FILES.get("archivo_autorizacion_padres")

        extensiones_permitidas = {
            ".pdf",
            ".doc",
            ".docx",
            ".jpg",
            ".jpeg",
            ".png",
        }

        for etiqueta, archivo in [
            ("documento de salud", archivo_salud),
            ("autorización de padres", archivo_autorizacion),
        ]:
            if not archivo:
                continue
            extension = Path(archivo.name).suffix.lower()
            if extension not in extensiones_permitidas:
                messages.error(
                    request,
                    f"El archivo de {etiqueta} no es válido. Solo se permiten PDF, Word o imagen.",
                )
                return render(
                    request,
                    "actualizar_info_asistente.html",
                    _contexto(),
                )

        documento_salud_requerido = None
        if archivo_salud:
            documento_salud_requerido = Documento.objects.filter(
                categoria="Formato Auto Reporte Condiciones de Salud"
            ).first()
            if not documento_salud_requerido:
                messages.error(
                    request,
                    "No se encontró el documento base de salud para registrar la actualización.",
                )
                return render(
                    request,
                    "actualizar_info_asistente.html",
                    _contexto(),
                )

        if not nombre or not tipo_doc or not numero_doc:
            messages.error(
                request, "Nombre, tipo y numero de documento son obligatorios."
            )
            return render(
                request,
                "actualizar_info_asistente.html",
                _contexto(),
            )

        duplicado_qs = asistente.__class__.objects.filter(
            visita=visita,
            numero_documento=numero_doc,
        ).exclude(id=asistente.id)

        if duplicado_qs.exists():
            messages.error(
                request,
                "Ya existe otro asistente en esta visita con ese numero de documento.",
            )
            return render(
                request,
                "actualizar_info_asistente.html",
                _contexto(),
            )

        asistente.nombre_completo = nombre
        asistente.tipo_documento = tipo_doc
        asistente.numero_documento = numero_doc
        asistente.correo = correo_asistente
        asistente.telefono = telefono
        campos_asistente_update = [
            "nombre_completo",
            "tipo_documento",
            "numero_documento",
            "correo",
            "telefono",
        ]

        if archivo_autorizacion:
            asistente.formato_autorizacion_padres = archivo_autorizacion
            asistente.estado_autorizacion_padres = "pendiente"
            asistente.observaciones_autorizacion_padres = ""
            campos_asistente_update.extend(
                [
                    "formato_autorizacion_padres",
                    "estado_autorizacion_padres",
                    "observaciones_autorizacion_padres",
                ]
            )

        try:
            aprendiz_a_actualizar = None
            campos_aprendiz_update = []

            if tipo == "interna":
                from panel_instructor_interno.models import Aprendiz, Ficha

                ficha = Ficha.objects.filter(numero=visita.numero_ficha).first()
                if ficha:
                    # Buscar el aprendiz vinculado por documento anterior para mantener el mismo registro.
                    aprendiz_a_actualizar = Aprendiz.objects.filter(
                        ficha=ficha, numero_documento=numero_doc_anterior
                    ).first()
                    if not aprendiz_a_actualizar:
                        aprendiz_a_actualizar = Aprendiz.objects.filter(
                            ficha=ficha, numero_documento=numero_doc
                        ).first()

                    if aprendiz_a_actualizar:
                        duplicado_aprendiz = Aprendiz.objects.filter(
                            ficha=ficha, numero_documento=numero_doc
                        ).exclude(id=aprendiz_a_actualizar.id)
                        if duplicado_aprendiz.exists():
                            messages.error(
                                request,
                                "No se puede actualizar porque ya existe otro aprendiz en la ficha con ese numero de documento.",
                            )
                            return render(
                                request,
                                "actualizar_info_asistente.html",
                                _contexto(),
                            )

                        partes_nombre = [p for p in nombre.split() if p]
                        apellido_actual_partes = [
                            p
                            for p in (aprendiz_a_actualizar.apellido or "").split()
                            if p
                        ]

                        # Intenta conservar apellidos compuestos usando la cantidad de palabras
                        # del apellido actual; si no hay referencia, asume 1 apellido.
                        if len(partes_nombre) <= 1:
                            nuevo_nombre = partes_nombre[0] if partes_nombre else nombre
                            nuevo_apellido = aprendiz_a_actualizar.apellido or "-"
                        else:
                            apellido_len = (
                                len(apellido_actual_partes)
                                if apellido_actual_partes
                                else 1
                            )
                            apellido_len = max(
                                1, min(apellido_len, len(partes_nombre) - 1)
                            )
                            nuevo_apellido = " ".join(partes_nombre[-apellido_len:])
                            nuevo_nombre = " ".join(partes_nombre[:-apellido_len])

                        aprendiz_a_actualizar.nombre = nuevo_nombre
                        aprendiz_a_actualizar.apellido = nuevo_apellido
                        aprendiz_a_actualizar.tipo_documento = tipo_doc
                        aprendiz_a_actualizar.numero_documento = numero_doc
                        aprendiz_a_actualizar.correo = correo_asistente
                        aprendiz_a_actualizar.telefono = telefono
                        campos_aprendiz_update = [
                            "nombre",
                            "apellido",
                            "tipo_documento",
                            "numero_documento",
                            "correo",
                            "telefono",
                            "fecha_actualizacion",
                        ]

                if not aprendiz_a_actualizar:
                    messages.warning(
                        request,
                        "Se actualizó el asistente, pero no se encontró un aprendiz asociado en la ficha para sincronizar.",
                    )

            with transaction.atomic():
                asistente.save(update_fields=campos_asistente_update)

                if archivo_salud:
                    DocumentoSubidoAsistente.objects.create(
                        documento_requerido=documento_salud_requerido,
                        asistente_interna=asistente if tipo == "interna" else None,
                        asistente_externa=asistente if tipo == "externa" else None,
                        archivo=archivo_salud,
                        estado="pendiente",
                        observaciones_revision="",
                    )

                if aprendiz_a_actualizar and campos_aprendiz_update:
                    if archivo_salud:
                        aprendiz_a_actualizar.documento_adicional = archivo_salud
                        if "documento_adicional" not in campos_aprendiz_update:
                            campos_aprendiz_update.append("documento_adicional")
                    aprendiz_a_actualizar.save(update_fields=campos_aprendiz_update)

            mensaje_exito = "Informacion del asistente actualizada correctamente." + (
                " También se sincronizó en la ficha del aprendiz."
                if tipo == "interna" and aprendiz_a_actualizar
                else ""
            )

            detalles_docs_actualizados = []
            if archivo_salud:
                detalles_docs_actualizados.append("documento de salud")
            if archivo_autorizacion:
                detalles_docs_actualizados.append("autorización de padres")

            if detalles_docs_actualizados:
                mensaje_exito += " Se actualizó: " + ", ".join(
                    detalles_docs_actualizados
                ) + "."

            if embed_mode:
                return render(
                    request,
                    "actualizar_info_asistente.html",
                    _contexto(
                        {
                            "actualizacion_exitosa": True,
                            "mensaje_exito": mensaje_exito,
                        }
                    ),
                )

            messages.success(request, mensaje_exito)
            return _redirect_segun_rol(request, tipo=tipo, visita_id=visita.id)
        except IntegrityError:
            messages.error(request, "No se pudo actualizar por conflicto de datos.")

    return render(
        request,
        "actualizar_info_asistente.html",
        _contexto(),
    )


def copiar_asistente_previo(request, tipo, visita_id, asistente_previo_id):
    """
    Copia un asistente previo de una visita anterior a la visita actual.
    """
    # Verificar autenticación
    if not request.session.get("responsable_autenticado"):
        return redirect("panel_visitante:login_responsable")

    correo = request.session.get("responsable_correo")
    # Obtener visita actual
    if tipo == "interna":
        visita = get_object_or_404(
            VisitaInterna,
            id=visita_id,
            correo_responsable__iexact=correo,
        )
        # Obtener asistente previo (de cualquier visita anterior con el mismo número de ficha)
        asistente_original = get_object_or_404(
            AsistenteVisitaInterna, id=asistente_previo_id, puede_reutilizar=True
        )
    elif tipo == "externa":
        visita = get_object_or_404(
            VisitaExterna,
            id=visita_id,
            correo_responsable__iexact=correo,
        )
        # Obtener asistente previo (de cualquier visita anterior con el mismo nombre)
        asistente_original = get_object_or_404(
            AsistenteVisitaExterna, id=asistente_previo_id, puede_reutilizar=True
        )
    else:
        messages.error(request, "Tipo de visita no válido.")
        return _redirect_segun_rol(request)

    # Verificar que no esté duplicado
    asistentes_actuales = visita.asistentes.all()
    max_asistentes = (
        visita.cantidad_aprendices if tipo == "interna" else visita.cantidad_visitantes
    )

    if asistentes_actuales.count() >= max_asistentes:
        messages.error(
            request, f"Ya se alcanzó el límite de {max_asistentes} asistentes."
        )
        return redirect(
            "panel_visitante:registrar_asistentes", tipo=tipo, visita_id=visita_id
        )

    # Verificar si el asistente ya está en la visita actual
    if asistentes_actuales.filter(
        numero_documento=asistente_original.numero_documento
    ).exists():
        messages.warning(
            request,
            f"El asistente {asistente_original.nombre_completo} ya está registrado en esta visita.",
        )
        return redirect(
            "panel_visitante:registrar_asistentes", tipo=tipo, visita_id=visita_id
        )

    try:
        # Copiar asistente
        if tipo == "interna":
            nuevo_asistente = AsistenteVisitaInterna.objects.create(
                visita=visita,
                nombre_completo=asistente_original.nombre_completo,
                tipo_documento=asistente_original.tipo_documento,
                numero_documento=asistente_original.numero_documento,
                correo=asistente_original.correo,
                telefono=asistente_original.telefono,
                es_reutilizado=True,
                visita_original=asistente_original,
            )
            # Copiar el archivo de autorización de padres si existe
            if asistente_original.formato_autorizacion_padres:
                from django.core.files.base import ContentFile

                archivo_original = asistente_original.formato_autorizacion_padres
                nuevo_asistente.formato_autorizacion_padres.save(
                    archivo_original.name,
                    ContentFile(archivo_original.read()),
                    save=True,
                )
        else:
            nuevo_asistente = AsistenteVisitaExterna.objects.create(
                visita=visita,
                nombre_completo=asistente_original.nombre_completo,
                tipo_documento=asistente_original.tipo_documento,
                numero_documento=asistente_original.numero_documento,
                correo=asistente_original.correo,
                telefono=asistente_original.telefono,
                es_reutilizado=True,
                visita_original=asistente_original,
            )
            # Copiar el archivo de autorización de padres si existe
            if asistente_original.formato_autorizacion_padres:
                from django.core.files.base import ContentFile

                archivo_original = asistente_original.formato_autorizacion_padres
                nuevo_asistente.formato_autorizacion_padres.save(
                    archivo_original.name,
                    ContentFile(archivo_original.read()),
                    save=True,
                )

        messages.success(
            request,
            f'Asistente "{nuevo_asistente.nombre_completo}" reutilizado correctamente de una visita anterior.',
        )
    except Exception as e:
        if "unique" in str(e).lower():
            messages.error(
                request,
                "Este asistente ya está registrado para esta visita.",
            )
        else:
            messages.error(request, f"Error al reutilizar: {str(e)}")

    return redirect(
        "panel_visitante:registrar_asistentes", tipo=tipo, visita_id=visita_id
    )
